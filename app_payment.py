import json
import time
import datetime
from psycopg2.extras import execute_values
from pprint import pprint
from flask import g, request, render_template, redirect, flash, url_for, session, abort, get_flashed_messages, \
    jsonify, Blueprint, current_app, send_file
from datetime import date, datetime
from flask_login import login_required
import error_handlers
import app_login
import pandas as pd
from openpyxl import Workbook
import os
import tempfile
import sys
import re

payment_app_bp = Blueprint('app_payment', __name__)

dbase = None

# Меню страницы
hlink_menu = None

# Меню профиля
hlink_profile = None


# Define a function to retrieve nonce within the application context
def get_nonce():
    with current_app.app_context():
        nonce = current_app.config.get('NONCE')
    return nonce


@payment_app_bp.before_request
def before_request():
    app_login.before_request()


# Проверка, что пользователь не уволен
@payment_app_bp.before_request
def check_user_status():
    app_login.check_user_status()


# Проверка, что ip адрес не забанен
@payment_app_bp.before_request
def is_ip_banned():
    app_login.is_ip_banned()

###################################################################################################
#               ДЛЯ ПО КОНТРОЛЬ ПЛАТЕЖЕЙ
###################################################################################################
@payment_app_bp.route('/po_payment_control')
def po_payment_control():
    try:
        if app_login.current_user.is_authenticated:
            app_login.set_info_log(log_url=sys._getframe().f_code.co_name, log_description=request.method,
                                   user_id=app_login.current_user.get_id(), ip_address=app_login.get_client_ip())
        else:
            app_login.set_info_log(log_url=sys._getframe().f_code.co_name, log_description=request.method,
                                   ip_address=app_login.get_client_ip())
        return render_template('__po_payment_control.html', nonce=get_nonce())
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())


@payment_app_bp.route('/download1pdf')
def download_file_1pdf():
    try:
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, ip_address=app_login.get_client_ip())
        path = 'static/po/1. Общее руководство.pdf'
        return send_file(path, as_attachment=True)
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return False


@payment_app_bp.route('/download2pdf')
def download_file_2pdf():
    try:
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, ip_address=app_login.get_client_ip())
        path = 'static/po/2. Инструкция по установке.pdf'
        return send_file(path, as_attachment=True)
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return False


@payment_app_bp.route('/download3pdf')
def download_file_3pdf():
    try:
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, ip_address=app_login.get_client_ip())
        path = 'static/po/3 Описание процессов жизненного цикла.pdf'
        return send_file(path, as_attachment=True)
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return False



@payment_app_bp.route('/payments', methods=["POST", "GET"])
@login_required
def payments():
    """Главная страница Платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        return render_template('payment-main.html', menu=hlink_menu, nonce=get_nonce(),
                               menu_profile=hlink_profile, title='Главная страница')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/new-payment', methods=['GET'])
@login_required
def get_new_payment():
    """Страница создания новой заявки на оплату"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init()

        # Список ответственных
        cursor.execute(
            "SELECT user_id, last_name, first_name FROM public.users WHERE is_fired = FALSE")
        responsible = cursor.fetchall()

        # Список типов заявок
        cursor.execute(
            """SELECT 
                    cost_item_id, 
                    cost_item_name, 
                    cost_item_category 
                FROM public.payment_cost_items 
                ORDER BY cost_item_category, cost_item_name""")
        cost_items_list = cursor.fetchall()

        # передаём данные в виде словаря для создания сгруппированного выпадающего списка
        cost_items_full = {}
        for item in cost_items_list:
            key = item[2]
            value = [item[1], item[0]]
            if key in cost_items_full:
                cost_items_full[key].append(value)
            else:
                cost_items_full[key] = [value]

        # Список объектов
        cursor.execute("SELECT object_id, object_name FROM public.objects")
        objects_name = cursor.fetchall()

        # Список контрагентов
        cursor.execute("SELECT DISTINCT partner FROM public.payments_summary_tab")
        partners = cursor.fetchall()

        # Get the current date
        today = date.today().strftime("%Y-%m-%d")

        # Список наших компаний из таблицы contractors
        cursor.execute("SELECT contractor_id, contractor_name FROM public.our_companies WHERE inflow_active")
        our_companies = cursor.fetchall()

        # Список типовых названий платежей пользователя
        cursor.execute(
            """
                SELECT 
                    DISTINCT SUBSTRING(basis_of_payment, 1,30) AS basis_of_payment
                FROM public.payments_summary_tab 
                WHERE payment_owner = %s OR responsible = %s
            """,
            [user_id, user_id]
        )
        bop = cursor.fetchall()

        # Close the database connection
        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()
        not_save_val = session['n_s_v_new_payment'] if session.get('n_s_v_new_payment') else {}

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        unselected = [int(x) for x in setting_users.keys()]

        cost_items = {}

        c_i_full_lst = []

        for k, v in cost_items_full.items():
            new_values = []
            for item in v:
                if item[1] not in unselected:
                    new_values.append(item.copy())
                    cost_items_full[k][cost_items_full[k].index(item)].append(1)
                c_i_full_lst.append(item)
            if new_values:
                cost_items[k] = new_values

        user_name = f'{app_login.current_user.get_last_name()} {app_login.current_user.get_name()}'

        return render_template('payment-new.html', responsible=responsible, cost_items=cost_items,
                               objects_name=objects_name, partners=partners, c_i_full_lst=c_i_full_lst,
                               our_companies=our_companies, menu=hlink_menu, menu_profile=hlink_profile,
                               user_id=user_id, user_name=user_name, nonce=get_nonce(),
                               bop=bop,
                               not_save_val=not_save_val, setting_users=setting_users, title='Новая заявка на оплату')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/new-payment', methods=['POST'])
@login_required
def set_new_payment():
    """Сохранение новой заявки на оплату в БД"""
    try:
        if request.method == 'POST':
            user_id = app_login.current_user.get_id()
            app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

            user_role_id = app_login.current_user.get_role()

            # Get the form data from the request
            basis_of_payment = request.form.get('basis_of_payment')  # Наименование платежа
            responsible = request.form.get('responsible').split('-@@@-')[0]  # Ответственный
            cost_items = request.form.get('cost_items').split('-@@@-')[1]  # Тип заявки
            try:
                object_id = request.form.get('objects_name').split('-@@@-')[0]  # id объекта
                object_name = request.form.get('objects_name').split('-@@@-')[-1]  # Название объекта
            except:
                object_id = None
                object_name = None
            payment_description = request.form.get('payment_description')  # Описание
            partner = request.form.get('partners')  # Контрагент
            payment_due_date = request.form.get('payment_due_date')  # Срок оплаты
            our_company_id = request.form.get('our_company').split('-@@@-')[0]  # id компании
            our_company = request.form.get('our_company').split('-@@@-')[1]  # Название компания
            payment_sum = request.form.get('payment_sum')  # Сумма оплаты
            payment_sum = convert_amount(payment_sum)
            payment_number = f'PAY-{round(time.time())}-___-{our_company}'  # Номера платежа

            # Connect to the database
            conn, cursor = app_login.conn_cursor_init()

            # Prepare the SQL query to insert the data into the payments_summary_tab
            query_s_t = """
            INSERT INTO public.payments_summary_tab (
                our_companies_id,
                cost_item_id,
                payment_number,
                basis_of_payment,
                payment_description,
                object_id,
                partner,
                payment_sum,
                payment_due_date,
                payment_owner,
                responsible
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            RETURNING payment_id;"""
            values_s_t = (
                our_company_id,
                cost_items,
                payment_number,
                basis_of_payment,
                payment_description,
                object_id,
                partner,
                payment_sum,
                payment_due_date,
                user_id,
                responsible)

            # Prepare the SQL query to insert the data into the payments_approval_history
            query_a_s = """
            INSERT INTO public.payments_approval_history (
                payment_id,
                status_id,
                user_id
            )
            VALUES (
                %s,
                %s,
                %s
            )"""

            try:
                """Запись в payments_summary_tab"""
                # Записываем новый платёж в БД и получаем обратно id записи для генерации номера платежа
                cursor.execute(query_s_t, values_s_t)
                last_payment_id = cursor.fetchone()[0]
                conn.commit()
                # Close the database connection
                app_login.conn_cursor_close(cursor, conn)

                # Execute the SQL query
                conn, cursor = app_login.conn_cursor_init()
                """Обновляем номер платежа в payments_summary_tab"""
                payment_number = f'PAY-{round(time.time())}-{last_payment_id}-{our_company}'
                query = """
                    UPDATE public.payments_summary_tab
                    SET payment_number = %s
                    WHERE payment_id = %s;
                """
                value = [payment_number, last_payment_id]
                cursor.execute(query, value)

                """Запись в payments_approval_history"""
                status_id_a_s = 1  # id статуса "Черновик"
                user_id_a_s = user_id if user_id else responsible
                values_a_s = (last_payment_id, status_id_a_s, user_id_a_s)
                cursor.execute(query_a_s, values_a_s)

                # Если создаёт бухгалтерия и рук., то добавляем статус "Реком." (создаётся доп. запись в бд)
                if user_role_id in (4, 6):
                    status_id_a_s = 5  # id статуса "Черновик"
                    values_a_s = (last_payment_id, status_id_a_s, user_id_a_s)
                    cursor.execute(query_a_s, values_a_s)

                conn.commit()

                # Close the database connection
                app_login.conn_cursor_close(cursor, conn)

                flash(message=['Платёж сохранён', f'№: {payment_number}'], category='success')
                session.pop('n_s_v_new_payment', default=None)

                return redirect(url_for('.get_new_payment'))
            except Exception as e:
                conn.rollback()
                # Close the database connection
                app_login.conn_cursor_close(cursor, conn)
                session['n_s_v_new_payment'] = {
                    'b_o_p': basis_of_payment,
                    'resp': [
                        responsible,
                        request.form.get('responsible').split('-@@@-')[-1]
                    ],
                    'c_i': [
                        request.form.get('cost_items').split('-@@@-')[0],
                        cost_items,
                        request.form.get('cost_items').split('-@@@-')[-1]
                    ],
                    'p_d': payment_description,
                    'part': partner,
                    'p_d_d': payment_due_date,
                    'o_c': [our_company_id, our_company],
                    'p_s': request.form.get('payment_sum')
                }
                if object_id:
                    session['n_s_v_new_payment']['obj_n'] = [object_id, object_name]

                msg_for_user = app_login.create_traceback(info=sys.exc_info())
                flash(message=['Платёж не сохранён', msg_for_user], category='error')
                return redirect(url_for('.get_new_payment'))
        return redirect(url_for('.get_new_payment'))

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info())
        flash(message=['Ошибка', msg_for_user], category='error')
        return redirect(url_for('.get_new_payment'))


@payment_app_bp.route('/payment-approval', methods=['GET'])
@login_required
def get_unapproved_payments():
    """Выгрузка из БД списка несогласованных платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        user_role_id = app_login.current_user.get_role()
        # Check if the user has access to the "List of contracts" page
        if user_role_id not in (1, 4, 6):
            return error_handlers.handle403(403)

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        page_name = 'payment-approval'

        cursor.execute(
            """
            SELECT 
                payment_id - 1 AS payment_id,
                (payment_due_date - interval '1 day')::date::text AS payment_due_date,
                payment_due_date::text AS initial_first_val,
                payment_id AS initial_id_val
            FROM public.payments_summary_tab
            WHERE not payment_close_status
            ORDER BY payment_due_date, payment_id
            LIMIT 1;
            """
        )
        all_payments = cursor.fetchall()

        # Список согласованных платежей
        cursor.execute("SELECT * FROM public.payments_approval")
        unapproved_payments = cursor.fetchall()

        # Список статусов платежей Андрея
        cursor.execute(
            """SELECT payment_agreed_status_id,
                      payment_agreed_status_name
            FROM public.payment_agreed_statuses WHERE payment_agreed_status_category = 'Andrew'""")
        approval_statuses = cursor.fetchall()

        # ДС на счету
        cursor.execute(
            """WITH
                t1 AS (SELECT 
                        COALESCE(sum(balance_sum), 0) AS account_money
                    FROM public.payments_balance),
                t2 AS (SELECT 
                        COALESCE(sum(approval_sum), 0) AS approval_sum
                    FROM public.payments_approval)
                SELECT 
                    t1.account_money AS account_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t1.account_money, 0), '999 999 990D99 ₽')) AS account_money_rub,
                    t1.account_money - t2.approval_sum AS available_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(COALESCE(t1.account_money - t2.approval_sum, t2.approval_sum), 0), '999 999 990D99 ₽')) AS available_money_rub,
                    t2.approval_sum AS approval_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t2.approval_sum, 0), '999 999 990D99 ₽')) AS approval_money_rub
                FROM t1
                JOIN t2 ON true;"""
        )
        money = cursor.fetchone()

        # Список ответственных
        cursor.execute(
            "SELECT user_id, last_name, first_name FROM public.users WHERE is_fired = FALSE ORDER BY last_name, first_name")
        responsible = cursor.fetchall()

        # Список типов заявок
        cursor.execute(
            """SELECT 
                cost_item_id, 
                cost_item_name, 
                cost_item_category 
            FROM public.payment_cost_items 
            ORDER BY cost_item_category, cost_item_name""")
        cost_items_list = cursor.fetchall()
        # передаём данные в виде словаря для создания сгруппированного выпадающего списка
        cost_items = {}
        for item in cost_items_list:
            key = item[2]
            value = [item[1], item[0]]
            if key in cost_items:
                cost_items[key].append(value)
            else:
                cost_items[key] = [value]

        # Список объектов
        cursor.execute("SELECT object_id, object_name FROM public.objects ORDER BY object_name")
        objects_name = cursor.fetchall()

        # Список контрагентов
        cursor.execute("SELECT DISTINCT partner FROM public.payments_summary_tab ORDER BY partner")
        partners = cursor.fetchall()

        # Список наших компаний из таблицы contractors
        cursor.execute("SELECT contractor_id, contractor_name FROM public.our_companies")
        our_companies = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [6, 0, all_payments[-1]['payment_due_date']],  # Первая колонка - ASC
                'col_id': all_payments[-1]['payment_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])
        tab_rows = 1
        # print('setting_users', setting_users)

        # Список колонок, которые скрываются для пользователя всегда
        hidden_col = []
        # print('user_role_id', user_role_id)
        if user_role_id == 6:
            # Для бухгалтерии срывается столбец "Согласованная сумма"
            hidden_col.append(4)

        # Список разделов в меню платежей
        payment_approval_menu = get_payment_approval_menu(user_role_id)
        if user_role_id in (1, 4, 6):
            payment_approval_menu[0]['class'] = 'focus_button'

        return render_template(
            'payment-approval.html', menu=hlink_menu, menu_profile=hlink_profile,
            # applications=all_payments,
            approval_statuses=approval_statuses, money=money, responsible=responsible,
            cost_items=cost_items, objects_name=objects_name, partners=partners, our_companies=our_companies,
            sort_col=sort_col, tab_rows=tab_rows, page=request.path[1:], setting_users=setting_users,
            hidden_col=hidden_col, nonce=get_nonce(), payment_approval_menu=payment_approval_menu,
            title='Согласование платежей')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-first-pay', methods=['POST'])
@login_required
def get_first_pay():
    """Постраничная выгрузка списка несогласованных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = request.get_json()['page_url']
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        if page_name == 'payment-approval-list' or page_name == 'payment-pay':
            col_id = 't0.payment_id'
        elif page_name == 'payment-inflow-history-list':
            col_id = 't1.inflow_at'
        else:
            col_id = 't1.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        if sort_col_1_order == 'DESC':
            order = '+'
        else:
            order = '-'

        if not where_expression2:
            where_expression2 = 'true'

        # pprint(request.get_json())

        if page_name == 'payment-approval':
            cursor.execute(
                f"""
                SELECT
                   t1.payment_id {order} 1 AS payment_id,
                   t5.first_name,
                   t5.last_name,
                   concat_ws(', ', t3.contractor_name, t6.object_name,
                   CASE
                       WHEN t1.partner<>'' THEN t1.partner
                   END) AS descr_part1,
                   concat_ws(' - ', t1.basis_of_payment, t1.payment_description) AS payment_description,
                   t1.payment_sum {order} 1 AS payment_sum,
                   COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum) {order} 1 AS approval_sum,
                   COALESCE(t8.amount, '0') {order} 1 AS amount,
                   (t1.payment_due_date {order} interval '1 day')::text AS payment_due_date,
                   (t1.payment_at {order} interval '1 day')::timestamp without time zone::text AS payment_at,
                   t7.status_name
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT
                            payment_id,
                            SUM(approval_sum) AS approval_sum
                        FROM public.payments_approval_history
                        GROUP BY payment_id
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            payment_id,
                            status_id
                        FROM public.payments_approval_history
                        WHERE status_id != 12
                        ORDER BY payment_id, confirm_id DESC
                ) AS t21 ON t1.payment_id = t21.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies
                ) AS t3 ON t1.our_companies_id = t3.contractor_id

                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT payment_agreed_status_id,
                            payment_agreed_status_name AS status_name
                        FROM public.payment_agreed_statuses
                ) AS t7 ON t21.status_id = t7.payment_agreed_status_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t1.payment_id = t8.payment_id
                WHERE not t1.payment_close_status AND {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #                 WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND {where_expression2}
            #                 ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #                 LIMIT {limit};""")
            # print('query_value', query_value)
        elif page_name == 'payment-approval-list':
            cursor.execute(
                f"""
                SELECT 
                    t0.payment_id {order} 1 AS payment_id,
                    t1.payment_number,
                    t3.contractor_name,
                    t1.basis_of_payment, 
                    t4.cost_item_name,
                    t1.payment_description, 
                    COALESCE(t6.object_name, ' ') AS object_name,
                    t5.first_name,
                    t5.last_name,
                    t1.partner,
                    t1.payment_sum {order} 1 AS payment_sum,
                    t0.approval_sum {order} 1 AS approval_sum,
                    COALESCE(t7.paid_sum, '0') {order} 1 AS paid_sum,
                    (t1.payment_due_date {order} interval '1 day')::text AS payment_due_date,
                    (t1.payment_at {order} interval '1 day')::timestamp without time zone::text AS payment_at,
                    (t8.created_at {order} interval '1 day')::timestamp without time zone::text AS created_at                    
                FROM public.payments_approval AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT 
                            DISTINCT payment_id,
                            SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                        FROM public.payments_paid_history
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                            SELECT DISTINCT ON (payment_id) 
                                payment_id,
                                created_at
                            FROM public.payments_approval_history
                            ORDER BY payment_id, created_at DESC
                    ) AS t8 ON t0.payment_id = t8.payment_id
                WHERE {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #     WHERE {where_expression2}
            #     ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #     LIMIT {limit};""")
            # print('query_value', query_value)
        elif page_name == 'payment-paid-list':
            cursor.execute(
                f"""
                WITH t0 AS (
                    SELECT 
                        payment_id,
                        MAX(created_at) AS paid_at,
                        SUM(paid_sum) AS paid_sum
                    FROM public.payments_paid_history
                    GROUP BY payment_id
                )
                SELECT 
                    t0.payment_id {order} 1 AS payment_id,
                    t1.payment_number,
                    t4.cost_item_name,
                    t1.basis_of_payment,
                    t3.contractor_name,
                    t1.payment_description,
                    COALESCE(t6.object_name, ' ') AS object_name,
                    t5.first_name,
                    t5.last_name,
                    t1.partner,
                    t1.payment_sum {order} 1 AS payment_sum,
                    COALESCE(t2.approval_sum, '0') {order} 1 AS approval_sum,
                    t0.paid_sum {order} 1 AS paid_sum,
                    (t1.payment_due_date {order} interval '1 day')::text AS payment_due_date,
                    (t1.payment_at {order} interval '1 day')::timestamp without time zone::text AS payment_at,
                    t8.status_name 
                FROM t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                        FROM public.payments_approval_history
                        ORDER BY payment_id, created_at DESC
                ) AS t2 ON t0.payment_id = t2.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            status_id
                        FROM public.payments_paid_history
                        ORDER BY payment_id, created_at DESC
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                        SELECT payment_agreed_status_id AS status_id,
                            payment_agreed_status_name AS status_name
                        FROM public.payment_agreed_statuses
                ) AS t8 ON t7.status_id = t8.status_id
                WHERE {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #     WHERE {where_expression2}
            #     ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #     LIMIT {limit};""")
            # print('query_value', query_value)
        elif page_name == 'payment-list':
            cursor.execute(
                f"""
                SELECT 
                    t1.payment_id {order} 1 AS payment_id,
                    t1.payment_number,
                    t4.cost_item_name,
                    t1.basis_of_payment,
                    t3.contractor_name,
                    t1.payment_description,
                    COALESCE(t6.object_name, ' ') AS object_name,
                    t5.first_name,
                    t5.last_name,
                    t1.partner,
                    t1.payment_sum {order} 1 AS payment_sum,
                    COALESCE(t7.paid_sum, 0) {order} 1 AS paid_sum,
                    (t1.payment_due_date {order} interval '1 day')::text AS payment_due_date,
                    t1.payment_at::timestamp without time zone::text AS payment_at
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            status_id,
                            SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                        FROM public.payments_approval_history
                        ORDER BY payment_id, created_at DESC
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                            SELECT 
                                DISTINCT payment_id,
                                SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                            FROM public.payments_paid_history
                    ) AS t7 ON t1.payment_id = t7.payment_id
                WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #     WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND {where_expression2}
            #     ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #     LIMIT {limit};""")
            # print('query_value', query_value)
        elif page_name == 'payment-pay':
            col_id = 't0.payment_id'
            cursor.execute(
                f"""
                SELECT 
                    t0.payment_id {order} 1 AS payment_id,
                    t4.cost_item_name,
                    t1.payment_number,
                    t1.basis_of_payment,
                    t1.payment_description, 
                    t3.contractor_name,
                    COALESCE(t6.object_name, ' ') AS object_name,
                    t5.first_name,
                    t5.last_name,
                    t1.partner,
                    t1.payment_sum {order} 1 AS payment_sum,
                    COALESCE(t7.paid_sum, '0') {order} 1 AS paid_sum,
                    t0.approval_sum {order} 1 AS approval_sum,
                    COALESCE(t8.amount, '0') {order} 1 AS amount,
                    (t1.payment_due_date {order} interval '1 day')::text AS payment_due_date,
                    (t1.payment_at {order} interval '1 day')::timestamp without time zone::text AS payment_at
                FROM public.payments_approval AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT 
                            DISTINCT payment_id,
                            SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                        FROM public.payments_paid_history
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t0.payment_id = t8.payment_id
                WHERE {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #     WHERE {where_expression2}
            #     ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #     LIMIT {limit};""")
            # print('query_value', query_value)
        elif page_name == 'payment-inflow-history-list':
            cursor.execute(
                f"""
                SELECT
                   t1.inflow_id {order} 1 AS inflow_id,
                   t2.contractor_name,
                   t1.inflow_description,
                   t3.inflow_type_name,
                   t1.inflow_sum {order} 1 AS inflow_sum,
                   (t1.inflow_at {order} interval '1 day')::timestamp without time zone::text AS inflow_at
                FROM public.payments_inflow_history AS t1
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t2 ON t1.inflow_company_id = t2.contractor_id
                LEFT JOIN (
                    SELECT inflow_type_id,
                        inflow_type_name
                    FROM public.payment_inflow_type
                ) AS t3 ON t1.inflow_type_id = t3.inflow_type_id
                WHERE {where_expression2}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # print(
            #     f"""/get-first-pay
            #                 WHERE {where_expression2}
            #                 ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
            #                 LIMIT {limit};""")
            # print('query_value', query_value)

        all_payments = cursor.fetchone()

        app_login.conn_cursor_close(cursor, conn)
        if all_payments:
            if page_name == 'payment-approval':
                col_0 = ""
                col_1 = f'{all_payments["descr_part1"]}: {all_payments["payment_description"]}'
                col_2 = all_payments["payment_sum"]
                col_3 = all_payments["approval_sum"]
                col_4 = all_payments["amount"]
                col_5 = f'{all_payments["last_name"]} {all_payments["first_name"]}'
                col_6 = all_payments["payment_due_date"]
                col_7 = all_payments["status_name"]
                col_8 = all_payments["payment_at"]
                col_9 = ""
                if sort_col_1_order == 'DESC':
                    col_1 = col_1 + '+'
                    col_5 = col_5 + '+'
                    col_7 = col_7 + '+'
                else:
                    col_1 = col_1[:-1]
                    col_5 = col_5[:-1]
                    col_7 = col_7[:-1]
                filter_col = [
                    col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
                ]
                sort_col['col_id'] = all_payments["payment_id"]
            elif page_name == 'payment-approval-list':
                col_0 = all_payments["payment_number"]
                col_1 = all_payments["cost_item_name"]
                col_2 = all_payments["basis_of_payment"]
                col_3 = f'{all_payments["contractor_name"]}: {all_payments["payment_description"]}'
                col_4 = all_payments["object_name"]
                col_5 = f'{all_payments["last_name"]} {all_payments["first_name"]}'
                col_6 = all_payments["partner"]
                col_7 = all_payments["payment_sum"]
                col_8 = all_payments["approval_sum"]
                col_9 = all_payments["paid_sum"]
                col_10 = all_payments["payment_due_date"]
                col_11 = all_payments["payment_at"]
                col_12 = all_payments["created_at"]
                if sort_col_1_order == 'DESC':
                    col_0 = col_0 + '+'
                    col_1 = col_1 + '+'
                    col_2 = col_2 + '+'
                    col_3 = col_3 + '+'
                    col_4 = col_4 + '+'
                    col_5 = col_5 + '+'
                    col_6 = col_6 + '+'
                else:
                    col_0 = col_0[:-1]
                    col_1 = col_1[:-1]
                    col_2 = col_2[:-1]
                    col_3 = col_3[:-1]
                    col_4 = col_4[:-1]
                    col_5 = col_5[:-1]
                    col_6 = col_6[:-1]
                filter_col = [
                    col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12
                ]
                sort_col['col_id'] = all_payments["payment_id"]
            elif page_name == 'payment-paid-list':
                col_0 = ""
                col_1 = all_payments["payment_number"]
                col_2 = all_payments["cost_item_name"]
                col_3 = all_payments["basis_of_payment"]
                col_4 = f'{all_payments["contractor_name"]} {all_payments["payment_description"]}'
                col_5 = all_payments["object_name"]
                col_6 = f'{all_payments["last_name"]} {all_payments["first_name"]}'
                col_7 = all_payments["partner"]
                col_8 = all_payments["payment_sum"]
                col_9 = all_payments["approval_sum"]
                col_10 = all_payments["paid_sum"]
                col_11 = all_payments["payment_due_date"]
                col_12 = all_payments["payment_at"]
                col_13 = all_payments["status_name"]

                if sort_col_1_order == 'DESC':
                    col_1 = col_1 + '+'
                    col_2 = col_2 + '+'
                    col_3 = col_3 + '+'
                    col_4 = col_4 + '+'
                    col_5 = col_5 + '='
                    col_6 = col_6 + '+'
                    col_7 = col_7 + '+'
                    col_13 = col_13 + '+'
                else:
                    col_1 = col_1[:-1]
                    col_2 = col_2[:-1]
                    col_3 = col_3[:-1]
                    col_4 = col_4[:-1]
                    col_5 = col_5[:-1]
                    col_6 = col_6[:-1]
                    col_7 = col_7[:-1]
                    col_13 = col_13[:-1]

                filter_col = [
                    col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
                ]
                sort_col['col_id'] = all_payments["payment_id"]
            elif page_name == 'payment-list':
                col_0 = all_payments["payment_number"]
                col_1 = all_payments["cost_item_name"]
                col_2 = all_payments["basis_of_payment"]
                col_3 = f'{all_payments["contractor_name"]} {all_payments["payment_description"]}'
                col_4 = all_payments["object_name"]
                col_5 = f'{all_payments["last_name"]} {all_payments["first_name"]}'
                col_6 = all_payments["partner"]
                col_7 = all_payments["payment_sum"]
                col_8 = all_payments["paid_sum"]
                col_9 = all_payments["payment_due_date"]
                col_10 = all_payments["payment_at"]
                if sort_col_1_order == 'DESC':
                    col_0 = col_0 + '+'
                    col_1 = col_1 + '+'
                    col_2 = col_2 + '+'
                    col_3 = col_3 + '+'
                    col_4 = col_4 + '='
                    col_5 = col_5 + '+'
                    col_6 = col_6 + '+'
                else:
                    col_0 = col_0[:-1]
                    col_1 = col_1[:-1]
                    col_2 = col_2[:-1]
                    col_3 = col_3[:-1]
                    col_4 = col_4[:-1]
                    col_5 = col_5[:-1]
                    col_6 = col_6[:-1]

                filter_col = [
                    col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10
                ]
                sort_col['col_id'] = all_payments["payment_id"]
            elif page_name == 'payment-pay':
                col_0 = ""
                col_1 = all_payments["cost_item_name"]
                col_2 = all_payments["payment_number"]
                col_3 = all_payments["basis_of_payment"]
                col_4 = f'{all_payments["contractor_name"]} {all_payments["payment_description"]}'
                col_5 = all_payments["object_name"]
                col_6 = f'{all_payments["last_name"]} {all_payments["first_name"]}'
                col_7 = all_payments["partner"]
                col_8 = all_payments["payment_sum"]
                col_9 = all_payments["paid_sum"]
                col_10 = all_payments["approval_sum"]
                col_11 = all_payments["amount"]
                col_12 = all_payments["payment_due_date"]
                col_13 = all_payments["payment_at"]
                col_14 = ""

                if sort_col_1_order == 'DESC':
                    col_1 = col_1 + '+'
                    col_2 = col_2 + '+'
                    col_3 = col_3 + '+'
                    col_4 = col_4 + '+'
                    col_5 = col_5 + '='
                    col_6 = col_6 + '+'
                    col_7 = col_7 + '+'
                else:
                    col_1 = col_1[:-1]
                    col_2 = col_2[:-1]
                    col_3 = col_3[:-1]
                    col_4 = col_4[:-1]
                    col_5 = col_5[:-1]
                    col_6 = col_6[:-1]
                    col_7 = col_7[:-1]

                filter_col = [
                    col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12,
                    col_13,
                    col_14
                ]
                sort_col['col_id'] = all_payments["payment_id"]
            elif page_name == 'payment-inflow-history-list':
                col_0 = all_payments["inflow_id"]
                col_1 = all_payments["contractor_name"]
                col_2 = all_payments["inflow_description"]
                col_3 = all_payments["inflow_type_name"]
                col_4 = all_payments["inflow_sum"]
                col_5 = all_payments["inflow_at"]
                if sort_col_1_order == 'DESC':
                    col_1 = col_1 + '+'
                    col_2 = col_2 + '+'
                    col_3 = col_3 + '+'
                else:
                    col_1 = col_1[:-1]
                    col_2 = col_2[:-1]
                    col_3 = col_3[:-1]
                filter_col = [col_0, col_1, col_2, col_3, col_4, col_5]

                sort_col['col_id'] = all_payments["inflow_id"]

            sort_col['col_1'].append(filter_col[col_num])

        # else:
        #     sort_col = {
        #         'col_1': [False, 0, False],  # Первая колонка
        #         'col_id': False
        #     }
        if not all_payments:
            return jsonify({
                'sort_col': sort_col,
                'status': 'error',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        return jsonify({
            'sort_col': sort_col,
            'status': 'success',
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/get-paymentApproval-pagination', methods=['POST'])
@login_required
def get_payment_approval_pagination():
    """Постраничная выгрузка списка несогласованных платежей"""

    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-approval'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't1.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        user_role_id = app_login.current_user.get_role()

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()
        try:
            cursor.execute(
                f"""
                SELECT
                    t1.payment_id,
                    t5.user_id,
                    t5.first_name,
                    t5.last_name,
                    concat_ws(', ', t3.contractor_name, t6.object_name,
                        CASE
                            WHEN t1.partner<>'' THEN t1.partner
                        END) AS descr_part1,
                    SUBSTRING(
                        concat_ws(' - ', t1.basis_of_payment, t1.payment_description),
                         1,70) AS payment_description_short,
                    concat_ws(' - ', t1.basis_of_payment, t1.payment_description) AS payment_description,
                    COALESCE(t6.object_name, '') AS object_name,
                    t1.partner,
                    t1.payment_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                    COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum) AS approval_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum), '999 999 990D99 ₽')) AS approval_sum_rub,
                    COALESCE(t8.amount, '0') AS amount,
                    COALESCE(t8.amount::text, '') AS amount_rub,
                    to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                    t1.payment_due_date::text AS payment_due_date,
                    t21.status_id,
                    to_char(payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt,
                    t1.payment_at::timestamp without time zone::text AS payment_at,
                    t1.payment_full_agreed_status,
                    t7.status_name
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT
                            payment_id,
                            SUM(approval_sum) AS approval_sum
                        FROM public.payments_approval_history
                        GROUP BY payment_id
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            payment_id,
                            status_id
                        FROM public.payments_approval_history
                        WHERE status_id != 12
                        ORDER BY payment_id, confirm_id DESC
                ) AS t21 ON t1.payment_id = t21.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies
                ) AS t3 ON t1.our_companies_id = t3.contractor_id

                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT payment_agreed_status_id,
                            payment_agreed_status_name AS status_name
                        FROM public.payment_agreed_statuses
                ) AS t7 ON t21.status_id = t7.payment_agreed_status_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t1.payment_id = t8.payment_id
                WHERE not t1.payment_close_status AND {where_expression}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            all_payments = cursor.fetchall()
            # print(cursor.query)

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = ""
        col_1 = f'{all_payments[-1]["descr_part1"]}: {all_payments[-1]["payment_description"]}'
        col_2 = all_payments[-1]["payment_sum"]
        col_3 = all_payments[-1]["approval_sum"]
        col_4 = all_payments[-1]["amount"]
        col_5 = f'{all_payments[-1]["last_name"]} {all_payments[-1]["first_name"]}'
        col_6 = all_payments[-1]["payment_due_date"]
        col_7 = all_payments[-1]["status_name"]
        col_8 = all_payments[-1]["payment_at"]
        col_9 = ""
        filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
        ]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["payment_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE not t1.payment_close_status AND ' + where_expression2
        else:
            where_expression2 = 'WHERE not t1.payment_close_status'

        # Число заявок
        cursor.execute(
            f"""SELECT
                    COUNT(t1.payment_id)
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT
                            payment_id,
                            SUM(approval_sum) AS approval_sum
                        FROM public.payments_approval_history
                        GROUP BY payment_id
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            payment_id,
                            status_id
                        FROM public.payments_approval_history
                        WHERE status_id != 12
                        ORDER BY payment_id, confirm_id DESC
                ) AS t21 ON t1.payment_id = t21.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies
                ) AS t3 ON t1.our_companies_id = t3.contractor_id

                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT payment_agreed_status_id,
                            payment_agreed_status_name AS status_name
                        FROM public.payment_agreed_statuses
                ) AS t7 ON t21.status_id = t7.payment_agreed_status_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id)
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t1.payment_id = t8.payment_id
                {where_expression2}
            """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        # Список статусов платежей Андрея
        cursor.execute(
            """SELECT payment_agreed_status_id,
                      payment_agreed_status_name
            FROM public.payment_agreed_statuses WHERE payment_agreed_status_category = 'Andrew'""")
        approval_statuses = cursor.fetchall()

        for i in range(len(approval_statuses)):
            approval_statuses[i] = dict(approval_statuses[i])

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'approval_statuses': approval_statuses,
            'user_role_id': user_role_id,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-approval', methods=['POST'])
@login_required
def set_approved_payments():
    """Сохранение согласованные платежи на оплату в БД"""
    try:
        user_id = app_login.current_user.get_id()

        if request.method == 'POST':
            app_login.set_info_log(log_url="payment-approval POST", user_id=user_id, ip_address=app_login.get_client_ip())

            # Ограничиваем доступ на изменение для бухгалтерии
            if app_login.current_user.get_role() not in (1, 4):
                flash(message=['Запрещено изменять данные', ''], category='error')
                app_login.set_fatal_error_log(log_url=sys._getframe().f_code.co_name,
                                              log_description='Ограничиваем доступ на изменение для бухгалтерии',
                                              user_id=user_id,
                                              ip_address=app_login.get_client_ip())
                return redirect(url_for('.get_unapproved_payments'))

            # Список выделенных столбцов
            selected_rows = request.form.getlist('selectedRows')  # Выбранные столбцы
            payment_number = request.form.getlist('payment_number')  # Номера платежей (передаётся id)
            status_id = request.form.getlist('status_id')  # Статус заявки (передаётся строковое название)
            payment_approval_sum = request.form.getlist('amount')  # Согласованная стоимость
            payment_full_agreed_status = request.form.getlist('payment_full_agreed_status')  # Сохранить до полной опл.

            selected_rows = [int(i) for i in selected_rows]
            payment_number = [int(i) for i in payment_number]

            payment_approval_sum = [convert_amount(i) for i in payment_approval_sum]
            payment_full_agreed_status = [int(i) for i in payment_full_agreed_status]

            values_p_s_t = []  # Данные для записи в таблицу payments_summary_tab
            values_p_a_h = []  # Данные для записи в таблицу payments_approval_history
            values_p_a = []  # Данные для записи в таблицу payments_approval_history

            # Данные для удаления временных данных из таблицы payments_summary_tab
            values_p_d = []
            page_name = 'payment-approval'
            parameter_name = 'amount'
            values_p_d_full_close = []  # Для полностью закрытых заявок удаляем черновики от всех пользователей

            values_a_h = []  # Список согласованных заявок для записи на БД
            pay_id_list_raw = []  # Список согласованных id заявок без обработки ошибок
            approval_id_list = []  # Список согласованных id заявок, без аннулир. и неправильные суммы согласования
            error_list = []  # Список id неправильно внесенных данных

            user_id = app_login.current_user.get_id()
            user_role_id = app_login.current_user.get_role()

            for i in selected_rows:
                row = i - 1

                pay_id_list_raw.append(payment_number[row])

                if user_role_id == 6:
                    payment_approval_sum[row] = ''

                # Проверка, что указана сумма согласования. Не для бухгалтера
                if user_role_id != 6 and not payment_approval_sum[row] and (
                        status_id[row] == 'Черновик' or status_id[row] == 'Реком.'):
                    flash(message=['Не указана сумма согласования', f'Платёж: id-{payment_number[row]}, строка: {i}'],
                          category='error')
                    return redirect(url_for('.get_unapproved_payments'))

                if status_id[row] == 'К рассмотрению':
                    flash(message=['Функция не работает', f'Платёж: id-{payment_number[row]}, строка: {i}'],
                          category='error')

                    return redirect(url_for('.get_unapproved_payments'))

                values_a_h.append([
                    payment_number[row],
                    status_id[row],
                    user_id,
                    payment_approval_sum[row]
                ])

            if not values_a_h:
                flash(message=['Не выбрано ни одной заявки', ''], category='error')
                return redirect(url_for('.get_unapproved_payments'))

            conn, cursor = app_login.conn_cursor_init_dict()

            cursor.execute(
                """WITH
                    t1 AS (SELECT 
                            COALESCE(sum(balance_sum), 0) AS account_money
                        FROM public.payments_balance),
                    t2 AS (SELECT 
                            COALESCE(sum(approval_sum), 0) AS approval_sum
                        FROM public.payments_approval)
                    SELECT 
                        (t1.account_money - t2.approval_sum)::float AS available_money
                    FROM t1
                    JOIN t2 ON true;"""
            )

            available_money = cursor.fetchone()[0]

            """
            Ищем сумму остатка согласования.
            - Если "согласованная сумма" равно остатку, то закрываем заявку, как полностью согласованную
            - Если не стоит галка "СОХРАНИТЬ ДО ПОЛНОЙ ОПЛАТЫ", полностью закрываем заявку со статусом "2"-
            Частичное согласование с закрытием
            """
            # Формируем обновленные данные для проверки:
            #  — id,
            #  — сумма оплаты,
            #  — ранее согласованное,
            #  — текущая сумма согл,
            #  — ранее согласованное + текущая сумма согл,
            #  — статус "Сохранить до полной опл",
            #  — статус закрытия,
            #  — статус Андрея
            query = """
                SELECT
                    t1.payment_id,
                    t1.payment_sum::float,
                    '' AS payment_approval_sum,
                    t2.approval_sum::float AS total_approval,
                    false AS payment_full_agreed_status,
                    false AS close_status,
                    'payment_agreed_statuses' AS status_id
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        sum (approval_sum) AS approval_sum
                    FROM public.payments_approval_history
                    GROUP BY payment_id
                ) AS t2 ON t1.payment_id = t2.payment_id
                WHERE t1.payment_id in %s;    
                """

            cursor.execute(query, (tuple(pay_id_list_raw),))
            total_approval_sum = cursor.fetchall()

            # Список статусов платежей Андрея
            cursor.execute(
                """SELECT payment_agreed_status_id AS id, 
                          payment_agreed_status_name  AS name
                FROM public.payment_agreed_statuses 
                WHERE payment_agreed_status_category = 'Andrew'
                """
            )
            approval_statuses = cursor.fetchall()

            # Добавляем недостающие данные и др. Подготавливаем данные для внесения в БД
            for i in range(len(total_approval_sum)):
                # Список id выбранных платежей
                for pay_id in pay_id_list_raw:
                    # Номер элемента в списках из формы
                    jj = payment_number.index(pay_id)
                    payment_approval_sum[jj] = payment_approval_sum[jj] if payment_approval_sum[jj] else 0

                    if total_approval_sum[i]['payment_id'] == pay_id:
                        # Сумма согласования
                        total_approval_sum[i]['payment_approval_sum'] = (
                            float(0 if payment_approval_sum[jj] is None else payment_approval_sum[jj]))

                        # Проверка, не выбрано ли бельше, чем можно согласовать
                        available_money -= total_approval_sum[i]['payment_approval_sum']
                        if available_money < 0:
                            flash(message=['Не хватает средств к распределению', ''], category='error')
                            app_login.conn_cursor_close(cursor, conn)
                            return redirect(url_for('.get_unapproved_payments'))

                        # ранее согласованное + текущая сумма согл
                        tot_app = float(0 if total_approval_sum[i]['total_approval'] is None
                                        else total_approval_sum[i]['total_approval'])
                        total_approval_sum[i]['total_approval'] = (
                                tot_app + total_approval_sum[i]['payment_approval_sum'])

                        # Статус "Сохранить до полной опл". Если галка стоит, то проставляем 1
                        for fas in payment_full_agreed_status:
                            if payment_number[fas - 1] == total_approval_sum[i]['payment_id']:
                                total_approval_sum[i]['payment_full_agreed_status'] = True
                                break

                        # Статус Андрея. Значения без учета сумм согласования
                        total_approval_sum[i]['status_id'] = status_id[jj]

                        # Статус закрытия и статус Андрея
                        # Не для бухгалтера
                        if user_role_id != 6:
                            if total_approval_sum[i]['status_id'] == 'Аннулирован':
                                total_approval_sum[i]['status_id'] = 6  # Полное согласование
                                total_approval_sum[i]['close_status'] = True  # Закрытие заявки
                            elif total_approval_sum[i][
                                'payment_full_agreed_status']:  # Если total_approval = payment_sum
                                if total_approval_sum[i]['payment_sum'] == total_approval_sum[i]['total_approval']:
                                    total_approval_sum[i]['status_id'] = 3  # Полное согласование
                                    total_approval_sum[i]['close_status'] = True  # Закрытие заявки
                            else:
                                total_approval_sum[i]['close_status'] = True  # Закрытие заявки

                                if total_approval_sum[i]['payment_sum'] != total_approval_sum[i]['total_approval']:
                                    if total_approval_sum[i]['payment_sum'] == total_approval_sum[i]['total_approval']:
                                        total_approval_sum[i]['status_id'] = 3  # Полное согласование
                                    else:
                                        total_approval_sum[i]['status_id'] = 2  # Частичное согласование с закрытием
                        # else:
                        #     if total_approval_sum[i]['status_id'] == 'Аннулирован':
                        #         total_approval_sum[i]['status_id'] = 6  # Полное согласование
                        #         total_approval_sum[i]['close_status'] = True  # Закрытие заявки
                        #     elif total_approval_sum[i]['status_id'] == 'Реком.':
                        #         total_approval_sum[i]['status_id'] = 5  # Полное согласование
                        #         total_approval_sum[i]['close_status'] = True  # Закрытие заявки

                # Переводим значение статуса Андрея из name в id
                for j2 in range(len(approval_statuses)):
                    if total_approval_sum[i]['status_id'] == approval_statuses[j2]['name']:
                        total_approval_sum[i]['status_id'] = approval_statuses[j2]['id']
                        # Если статус Аннулировано (id 6), то закрываем заявку
                        if total_approval_sum[i]['status_id'] == 6:
                            total_approval_sum[i]['close_status'] = True  # Закрытие заявки

                # Проверка, что общая согласованная сумма меньше либо равна сумме к оплате
                if total_approval_sum[i]['total_approval'] > total_approval_sum[i]['payment_sum']:
                    error_list.append([
                        total_approval_sum[i]['payment_id'],
                        total_approval_sum[i]['payment_sum'],
                        total_approval_sum[i]['total_approval'],
                        total_approval_sum[i]['payment_approval_sum'],
                        'Общая сумма согласования больше остатка'

                    ])
                    total_approval_sum[i]['payment_id'] = None

            # Создаём списки с данными для записи в БД
            for i in range(len(total_approval_sum)):
                """для db payments_summary_tab"""
                if total_approval_sum[i]['payment_id']:
                    values_p_s_t.append([
                        total_approval_sum[i]['payment_id'],
                        total_approval_sum[i]['payment_full_agreed_status'],
                        total_approval_sum[i]['close_status']
                    ])
                    values_p_d.append((
                        page_name,
                        total_approval_sum[i]['payment_id'],
                        parameter_name,
                        user_id
                    ))
                    if total_approval_sum[i]['close_status']:
                        values_p_d_full_close.append((
                            page_name,
                            total_approval_sum[i]['payment_id']
                        ))

                """для db payments_approval_history"""
                if total_approval_sum[i]['payment_id'] and total_approval_sum[i]['status_id'] in [2, 3, 4, 5, 6]:
                    values_p_a_h.append([
                        total_approval_sum[i]['payment_id'],
                        total_approval_sum[i]['status_id'],
                        user_id,
                        total_approval_sum[i]['payment_approval_sum']
                    ])

                """для db payments_approva"""
                # Если есть id заявки и не Аннулировано (status_id 6). Не для Бухгалтера
                if (user_role_id != 6 and total_approval_sum[i]['payment_id'] and
                        total_approval_sum[i]['status_id'] != 6):
                    values_p_a.append([
                        total_approval_sum[i]['payment_approval_sum'],
                        total_approval_sum[i]['payment_id']
                    ])

            try:
                # Если есть что записывать в Базу данных
                if values_p_s_t:
                    # Перезапись в payments_summary_tab
                    columns_p_s_t = ("payment_id", "payment_full_agreed_status", "payment_close_status")
                    query_p_s_t = get_db_dml_query(action='UPDATE', table='payments_summary_tab', columns=columns_p_s_t)
                    execute_values(cursor, query_p_s_t, values_p_s_t)

                    columns_p_d = 'page_name, parent_id::int, parameter_name, user_id'
                    query_p_d = get_db_dml_query(action='DELETE', table='payment_draft', columns=columns_p_d)
                    execute_values(cursor, query_p_d, (values_p_d,))

                    if values_p_d_full_close:
                        columns_p_d_full_close = 'page_name, parent_id::int'
                        query_p_d_full_close = get_db_dml_query(action='DELETE', table='payment_draft',
                                                                columns=columns_p_d_full_close)
                        execute_values(cursor, query_p_d_full_close, (values_p_d_full_close,))

                    # Если есть что записывать в payments_approval_history
                    if values_p_a_h:
                        # Запись в payments_approval_history
                        action_p_a_h = 'INSERT INTO'
                        table_p_a_h = 'payments_approval_history'
                        columns_p_a_h = ('payment_id', 'status_id', 'user_id', 'approval_sum')
                        subquery = " RETURNING payment_id, confirm_id;"
                        query_a_h = get_db_dml_query(
                            action=action_p_a_h, table=table_p_a_h, columns=columns_p_a_h, subquery=subquery
                        )
                        execute_values(cursor, query_a_h, values_p_a_h)

                    # Если есть что записывать в payments_approval_history. Не для Бухгалтера
                    if user_role_id != 6 and values_p_a:
                        # # Запись в payments_approval
                        action_p_a = 'INSERT CONFLICT UPDATE'
                        table_p_a = 'payments_approval'
                        columns_p_a = ('approval_sum', 'payment_id')

                        expr_set = ', '.join([f"{col} = t1.{col} + EXCLUDED.{col}" for col in columns_p_a[:-1]])
                        query_p_a = get_db_dml_query(
                            action=action_p_a, table=table_p_a, columns=columns_p_a, expr_set=expr_set
                        )

                        execute_values(cursor, query_p_a, values_p_a)

                    conn.commit()

                    flash(message=['Заявки согласованы', ''], category='success')

                else:
                    flash(message=['Нет данных для сохранения', ''], category='error')
                # Если есть ошибки
                if error_list:
                    flash(message=[error_list, ''], category='error')

                app_login.conn_cursor_close(cursor, conn)

                return redirect(url_for('.get_unapproved_payments'))
            except Exception as e:
                conn.rollback()
                app_login.conn_cursor_close(cursor, conn)
                msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
                return redirect(url_for('.get_unapproved_payments'))

        return redirect(url_for('.get_unapproved_payments'))

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return redirect(url_for('.get_unapproved_payments'))


@payment_app_bp.route('/save_quick_changes_approved_payments', methods=['POST'])
@login_required
def save_quick_changes_approved_payments():
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Сохраняем изменения в полях (согл сумма, статус, сохр до полн оплаты) заявки без нажатия кнопки "Отправить"
        page = request.form['page']
        payment_id = int(request.form['payment_number'])
        row_id = request.form['row_id']
        amount = convert_amount(request.form['amount'])

        if page == 'payment-approval':
            status_id = request.form['status_id']
            status_id2 = request.form.getlist('status_id')
        else:
            status_id = None
            status_id2 = None
        agreed_status = request.form['payment_full_agreed_status']
        # Преобразовываем в нужный тип данных
        if agreed_status == 'false':
            agreed_status = False
        else:
            agreed_status = True
        if amount:
            amount = float(amount)

        # Execute the SQL query
        conn, cursor = app_login.conn_cursor_init()

        if page == 'payment-approval':
            # Статусы Андрея
            query_approval_statuses = """
                SELECT payment_agreed_status_id AS id
                FROM public.payment_agreed_statuses 
                WHERE payment_agreed_status_name = %s
            """
            cursor.execute(query_approval_statuses, (status_id,))
            approval_statuses = cursor.fetchone()[0]

            # СТАТУС ПЛАТЕЖА
            # Последний статус платежа
            query_last_status = """
                SELECT DISTINCT ON (payment_id) 
                         status_id
                FROM public.payments_approval_history
                WHERE payment_id = %s
                ORDER BY payment_id, created_at DESC
            """
            cursor.execute(query_last_status, (payment_id,))
            try:
                last_status_id = cursor.fetchone()[0]
            except:
                last_status_id = ''
            # Если статус New (id 1), приравниваем его к id 4 - Черновик
            if last_status_id == 1:
                last_status_id = 4
            # Если статусы не совпадают, создаём новую запись
            if last_status_id != approval_statuses:
                # Запись в payments_approval_history
                action_p_a_h = 'INSERT INTO'
                table_p_a_h = 'payments_approval_history'
                columns_p_a_h = ('payment_id', 'status_id', 'user_id')
                values_p_a_h = [[payment_id, approval_statuses, user_id]]
                query_a_h = get_db_dml_query(action=action_p_a_h, table=table_p_a_h, columns=columns_p_a_h)
                execute_values(cursor, query_a_h, values_p_a_h)

            # СОХРАНИТЬ ДО ПОЛНОЙ ОПЛАТЫ
            # Последний статус сохранения до полной оплаты
            query_last_f_a_status = """
                SELECT payment_full_agreed_status
                FROM public.payments_summary_tab
                WHERE payment_id = %s
            """
            cursor.execute(query_last_f_a_status, (payment_id,))
            last_f_a_status = cursor.fetchone()[0]
            # Если статусы не совпадают, обновляем запись
            if last_f_a_status != agreed_status:
                columns_p_s_t = ("payment_id", "payment_full_agreed_status")
                values_p_s_t = [[payment_id, agreed_status]]
                query_p_s_t = get_db_dml_query(action='UPDATE', table='payments_summary_tab', columns=columns_p_s_t)
                execute_values(cursor, query_p_s_t, values_p_s_t)

        if page == 'payment-pay':
            # ЗАКРЫТЬ ТОЛЬКО ПОСЛЕ ПОЛНОЙ ОПЛАТЫ
            columns_p_a = ("payment_id", "approval_fullpay_close_status")
            values_p_a = [[payment_id, agreed_status]]
            query_p_a = get_db_dml_query(action='UPDATE', table='payments_approval', columns=columns_p_a)
            execute_values(cursor, query_p_a, values_p_a)

        # СОГЛАСОВАННАЯ СУММА
        # Неотправленная согласованная сумма
        query_last_amount = """
        SELECT DISTINCT ON (payment_id) 
            parent_id::int AS payment_id,
            parameter_value::float AS amount
        FROM public.payment_draft
        WHERE page_name = %s AND parent_id::int = %s AND parameter_name = %s AND user_id = %s
        ORDER BY payment_id, created_at DESC;
        """
        parameter_name = 'amount'
        value_last_amount = [page, payment_id, parameter_name, user_id]
        cursor.execute(query_last_amount, value_last_amount)
        last_amount = cursor.fetchone()
        if last_amount:
            last_amount = last_amount[1]
        # Если суммы не совпадают, добавляем запись
        if amount != last_amount:
            # Если неотправленна сумма была, то удаляем её и вносим новую (удаляем, а не перезаписываем т.к.
            # возможно в таблице может быть несколько записей)
            if last_amount:
                # Удаление всех неотправленных сумм
                cursor.execute("""
                DELETE FROM public.payment_draft 
                WHERE page_name = %s AND parent_id::int = %s AND parameter_name = %s AND user_id = %s
                """, value_last_amount)
            # Если указали сумму согласования, то вносим в таблицу временных значений, иначе не вносим
            if amount:
                action_d_p = 'INSERT INTO'
                table_d_p = 'payment_draft'
                columns_d_p = ('page_name', 'parent_id', 'parameter_name', 'parameter_value', 'user_id')
                values_d_p = [[page, payment_id, parameter_name, amount, user_id]]
                query_d_p = get_db_dml_query(action=action_d_p, table=table_d_p, columns=columns_d_p)
                execute_values(cursor, query_d_p, values_d_p)

        conn.commit()

        app_login.conn_cursor_close(cursor, conn)

        return 'Data saved successfully'
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return msg_for_user


@payment_app_bp.route('/cash-inflow', methods=['GET'])
@login_required
def get_cash_inflow():
    """Страница для добавления платежа"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Check if the user has access to the "List of contracts" page
        if app_login.current_user.get_role() not in (1, 6):
            return error_handlers.handle403(403)

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        # Список наших компаний из таблицы contractors
        cursor.execute(
            "SELECT contractor_id, contractor_name FROM public.our_companies WHERE inflow_active is true"
        )
        our_companies = cursor.fetchall()

        # Список типов поступлений из таблицы payment_inflow_type
        cursor.execute("SELECT * FROM public.payment_inflow_type")
        inflow_types = cursor.fetchall()

        # Последние 5 поступлений из таблицы payment_inflow_type
        cursor.execute("""
        SELECT 
            date_trunc('second', t1.inflow_at::timestamp without time zone)::text AS inflow_at,
            TRIM(BOTH ' ' FROM to_char(inflow_sum, '999 999 990D99 ₽')) AS inflow_sum,
            t2.contractor_name,
            t1.inflow_description            
        FROM public.payments_inflow_history AS t1
        LEFT JOIN (
                    SELECT  
                        contractor_id,
                        contractor_name
                    FROM public.our_companies
            ) AS t2 ON t1.inflow_company_id = t2.contractor_id
        ORDER BY inflow_at DESC LIMIT 5""")
        historical_data = cursor.fetchall()

        # Список балансов компаний
        cursor.execute("""
        SELECT 
            t1.contractor_name,
            TRIM(BOTH ' ' FROM to_char(COALESCE(t2.balance_sum, 0), '999 999 990D99 ₽')) AS balance_sum
        FROM public.our_companies AS t1
        LEFT JOIN (
                    SELECT  
                        company_id,
                        balance_sum
                    FROM public.payments_balance
            ) AS t2 ON t1.contractor_id = t2.company_id
        ORDER BY t1.contractor_id 
        LIMIT 3
        """)
        companies_balances = cursor.fetchall()

        # Список балансов других компаний
        cursor.execute("""
        SELECT 
            t1.contractor_name,
            TRIM(BOTH ' ' FROM to_char(COALESCE(t2.balance_sum, 0), '999 999 990D99 ₽')) AS balance_sum        
        FROM public.our_companies AS t1
        LEFT JOIN (
                    SELECT  
                        company_id,
                        balance_sum
                    FROM public.payments_balance
            ) AS t2 ON t1.contractor_id = t2.company_id
        WHERE t1.inflow_active
        ORDER BY t1.contractor_id 
        OFFSET 3
        """)
        subcompanies_balances = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()
        not_save_val = session['n_s_v_cash_inflow'] if session.get('n_s_v_cash_inflow') else {}

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        return render_template(
            template_name_or_list='payment-cash-inflow.html', menu=hlink_menu, menu_profile=hlink_profile,
            our_companies=our_companies, inflow_types=inflow_types, historical_data=historical_data,
            not_save_val=not_save_val, companies_balances=companies_balances, page=request.path[1:],
            subcompanies_balances=subcompanies_balances, setting_users=setting_users, nonce=get_nonce(),
            title='Поступления денежных средств')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/cash-inflow', methods=['POST'])
@login_required
def set_cash_inflow():
    """Сохранение согласованные платежи на оплату в БД"""
    try:
        if request.method == 'POST':
            user_id = app_login.current_user.get_id()
            app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

            # Список выделенных столбцов
            inflow_company_id = int(request.form.get('company_ci').split('-@@@-')[0])  # id компании
            inflow_company = request.form.get('company_ci').split('-@@@-')[1]  # Название компания
            inflow_type_id = int(request.form.get('type_ci').split('-@@@-')[0])  # id типа поступления
            inflow_type = request.form.get('type_ci').split('-@@@-')[1]  # Название типа поступления
            inflow_sum = convert_amount(request.form['money_ci'])  # Сумма поступления
            try:
                transfer_company_id = int(request.form.get('transfer_company_ci').split('-@@@-')[0])  # id компании
                transfer_company = request.form.get('transfer_company_ci').split('-@@@-')[1]  # Название компания
            except:
                transfer_company_id = None
                transfer_company = None
            try:
                inflow_description = request.form['description_ci']  # Комментарий
            except:
                inflow_description = None

            action_i_h = 'INSERT INTO'
            table_i_h = 'payments_inflow_history'
            columns_i_h = ('inflow_company_id', 'inflow_description', 'inflow_type_id', 'inflow_sum', 'inflow_owner')
            query_i_h = get_db_dml_query(action=action_i_h, table=table_i_h, columns=columns_i_h)
            values_i_h = [[inflow_company_id, inflow_description, inflow_type_id, inflow_sum, user_id]]

            action_b = 'INSERT CONFLICT UPDATE'
            table_b = 'payments_balance'
            columns_b = ('balance_sum', 'company_id')

            conn, cursor = app_login.conn_cursor_init()

            try:
                # Если Тип поступления "Поступление ДС",
                # то добавляем данные в таблицы payments_inflow_history и payments_balance
                if inflow_type_id == 1:
                    # Запись в таблицу payments_inflow_history
                    execute_values(cursor, query_i_h, values_i_h)

                    # Запись в таблицу payments_balance
                    # Генерируем выражение: к текущему значению всех колонок добавляем новое
                    expr_set = ', '.join([f"{col} = t1.{col} + EXCLUDED.{col}" for col in columns_b[:-1]])

                    query_b = get_db_dml_query(action=action_b, table=table_b, columns=columns_b, expr_set=expr_set)
                    values_b = [[inflow_sum, inflow_company_id]]
                    execute_values(cursor, query_b, values_b)
                    flash(message=['Поступление добавлено', ''], category='success')

                # Если Тип поступления "П.О.", то пока ничего не делаем
                elif inflow_type_id == 2:
                    flash(message=['Действие отменено', 'Тип поступления П.О. не работает'], category='error')
                    return redirect(url_for('.get_cash_inflow'))

                # Если Тип поступления "Корректирующий платеж",
                # то перемещаем средства между компания в таблице payments_balance
                elif inflow_type_id == 3:
                    flash(message=['Действие отменено', 'Тип поступления \"Корректирующий платеж\" не работает'],
                          category='error')
                    return redirect(url_for('.get_cash_inflow'))

                # Если Тип поступления "Внутренний платеж",
                # то перемещаем средства между компания в таблице payments_balance
                elif inflow_type_id == 4:
                    # Проверяем, хватает ли средств у inflow_company
                    query = """
                    SELECT 
                        balance_sum 
                    FROM public.payments_balance 
                    WHERE company_id::int = %s
                    """
                    execute_values(cursor, query, [[inflow_company_id]])
                    balance_sum = cursor.fetchone()
                    balance_sum = 0 if not balance_sum else balance_sum[0]
                    if balance_sum < inflow_sum:
                        flash(message=[
                            'Действие отменено',
                            f'На счету компании: {inflow_company} недостаточно средств ({balance_sum} ₽) '
                            f'для перевода.\nНе хватает:  {inflow_sum - float(balance_sum)} ₽\n\nОтмена операции'],
                            category='error')
                        return redirect(url_for('.get_cash_inflow'))

                    # Запись в таблицу payments_inflow_history
                    inflow_description = f"из {inflow_company} {inflow_sum} ₽"
                    query_i_h = get_db_dml_query(action=action_i_h, table=table_i_h, columns=columns_i_h)
                    values_i_h = [[transfer_company_id, inflow_description, inflow_type_id, inflow_sum, user_id]]
                    execute_values(cursor, query_i_h, values_i_h)

                    # Запись в таблицу payments_balance
                    # Генерируем выражение: к текущему значению всех колонок добавляем новое. Прибавляем у taker_comp
                    expr_set = ', '.join([f"{col} = t1.{col} + EXCLUDED.{col}" for col in columns_b[:-1]])
                    query_b = get_db_dml_query(action=action_b, table=table_b, columns=columns_b, expr_set=expr_set)
                    values_b = [[inflow_sum, transfer_company_id]]
                    execute_values(cursor, query_b, values_b)
                    # Генерируем выражение: из тек. знач. вычитаем (прибалвяем отрицательную inflow_sum. Вычитание у inflow_company
                    expr_set = ', '.join([f"{col} = t1.{col} + EXCLUDED.{col}" for col in columns_b[:-1]])
                    query_b = get_db_dml_query(action=action_b, table=table_b, columns=columns_b, expr_set=expr_set)
                    # inflow_sum = -inflow_sum
                    values_b = [[-inflow_sum, inflow_company_id]]
                    execute_values(cursor, query_b, values_b)

                    flash(message=['Внутренний платеж осуществлён', ''], category='success')

                conn.commit()

                app_login.conn_cursor_close(cursor, conn)
                session.pop('n_s_v_cash_inflow', default=None)

                return redirect(url_for('.get_cash_inflow'))

            except Exception as e:
                conn.rollback()
                app_login.conn_cursor_close(cursor, conn)

                session['n_s_v_cash_inflow'] = {
                    'o_c': [inflow_company_id, inflow_company],
                    'i_t': [inflow_type_id, inflow_type],
                    'c_i_s': request.form.get('money_ci'),
                }
                if transfer_company_id:
                    session['n_s_v_cash_inflow']['t_c'] = [transfer_company_id, transfer_company]
                if inflow_description:
                    session['n_s_v_cash_inflow']['i_d'] = inflow_description

                msg_for_user = app_login.create_traceback(info=sys.exc_info())
                flash(message=['Ошибка. Данные не сохранены', msg_for_user], category='error')

                return redirect(url_for('.get_cash_inflow'))

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return redirect(url_for('.get_cash_inflow'))


@payment_app_bp.route('/payment-pay', methods=['GET'])
@login_required
def get_unpaid_payments():
    """Выгрузка из БД списка неоплаченных платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Check if the user has access to the "List of contracts" page
        if app_login.current_user.get_role() not in (1, 6):
            return error_handlers.handle403(403)

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        page_name = 'payment-pay'

        cursor.execute(
            f"""
            SELECT 
                t0.payment_id - 1 AS payment_id,
                (t1.payment_due_date - interval '1 day')::date::text AS payment_due_date
            FROM public.payments_approval AS t0
            LEFT JOIN (
                SELECT 
                    payment_id,
                    payment_due_date
                FROM public.payments_summary_tab
            ) AS t1 ON t0.payment_id = t1.payment_id
            ORDER BY t1.payment_due_date, t0.payment_id
            LIMIT 1;
            """
        )
        all_payments = cursor.fetchall()

        # Список статусов платежей Андрея
        cursor.execute(
            """SELECT payment_agreed_status_id,
                      payment_agreed_status_name
            FROM public.payment_agreed_statuses WHERE payment_agreed_status_category = 'Andrew'""")
        approval_statuses = cursor.fetchall()

        # ДС на счету
        cursor.execute(
            """WITH
                t1 AS (SELECT 
                        COALESCE(SUM(balance_sum), 0) AS account_money
                    FROM public.payments_balance),
                t2 AS (SELECT 
                        COALESCE(SUM(approval_sum), 0) AS approval_sum
                    FROM public.payments_approval)
                SELECT 
                    t1.account_money AS account_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t1.account_money, 0), '999 999 990D99 ₽')) AS account_money_rub,
                    t1.account_money - t2.approval_sum AS available_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(COALESCE(t1.account_money - t2.approval_sum, t2.approval_sum), 0), '999 999 990D99 ₽')) AS available_money_rub
                FROM t1
                JOIN t2 ON true;"""
        )
        money = cursor.fetchone()

        # Список ответственных
        cursor.execute(
            "SELECT user_id, last_name, first_name FROM public.users WHERE is_fired = FALSE ORDER BY last_name, first_name")
        responsible = cursor.fetchall()

        # Список типов заявок
        cursor.execute(
            """SELECT 
                cost_item_id, 
                cost_item_name, 
                cost_item_category 
            FROM public.payment_cost_items 
            ORDER BY cost_item_category, cost_item_name""")
        cost_items_list = cursor.fetchall()
        # передаём данные в виде словаря для создания сгруппированного выпадающего списка
        cost_items = {}
        for item in cost_items_list:
            key = item[2]
            value = [item[1], item[0]]
            if key in cost_items:
                cost_items[key].append(value)
            else:
                cost_items[key] = [value]

        # Список объектов
        cursor.execute("SELECT object_id, object_name FROM public.objects ORDER BY object_name")
        objects_name = cursor.fetchall()

        # Список контрагентов
        cursor.execute("SELECT DISTINCT partner FROM public.payments_summary_tab ORDER BY partner")
        partners = cursor.fetchall()

        # Список наших компаний из таблицы contractors
        cursor.execute("SELECT contractor_id, contractor_name FROM public.our_companies")
        our_companies = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [12, 0, all_payments[-1]['payment_due_date']],  # Первая колонка - ASC
                'col_id': all_payments[-1]['payment_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])
        tab_rows = 1

        return render_template(
            'payment-pay.html', menu=hlink_menu, menu_profile=hlink_profile, nonce=get_nonce(),
            # applications=all_payments,
            responsible=responsible, cost_items=cost_items, objects_name=objects_name,
            partners=partners, our_companies=our_companies,
            approval_statuses=approval_statuses, money=money,
            sort_col=sort_col, tab_rows=tab_rows, page=request.path[1:], setting_users=setting_users,
            title='Оплата платежей')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-paymentPay-pagination', methods=['POST'])
@login_required
def get_payment_pay_pagination():
    """Постраничная выгрузка списка согласованных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-pay'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't0.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            cursor.execute(
                f"""
                SELECT 
                    t0.payment_id,
                    t1.payment_number,
                    t3.contractor_name,
                    t3.contractor_id,
                    t4.cost_item_name,
                    SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,
                    t1.basis_of_payment,  
                    t5.first_name,
                    t5.last_name,
                    SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                    t1.payment_description, 
                    COALESCE(t6.object_name, ' ') AS object_name,
                    t1.partner,
                    t1.payment_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                    COALESCE(t7.paid_sum, '0') AS paid_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t7.paid_sum, 0), '999 999 990D99 ₽')) AS paid_sum_rub,
                    t0.approval_sum,
                    TRIM(BOTH ' ' FROM to_char(t0.approval_sum, '999 999 990D99 ₽')) AS approval_sum_rub,
                    COALESCE(t8.amount, '0') AS amount,
                    COALESCE(t8.amount::text, '') AS amount_rub,
                    to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                    t1.payment_due_date::text AS payment_due_date,
                    t0.approval_fullpay_close_status AS payment_full_agreed_status,
                    to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt,
                    t1.payment_at::timestamp without time zone::text AS payment_at
                FROM public.payments_approval AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT 
                            DISTINCT payment_id,
                            SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                        FROM public.payments_paid_history
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t0.payment_id = t8.payment_id
                WHERE {where_expression}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = ""
        col_1 = all_payments[-1]["cost_item_name"]
        col_2 = all_payments[-1]["payment_number"]
        col_3 = all_payments[-1]["basis_of_payment"]
        col_4 = f'{all_payments[-1]["contractor_name"]} {all_payments[-1]["payment_description"]}'
        col_5 = all_payments[-1]["object_name"]
        col_6 = f'{all_payments[-1]["last_name"]} {all_payments[-1]["first_name"]}'
        col_7 = all_payments[-1]["partner"]
        col_8 = all_payments[-1]["payment_sum"]
        col_9 = all_payments[-1]["paid_sum"]
        col_10 = all_payments[-1]["approval_sum"]
        col_11 = all_payments[-1]["amount"]
        col_12 = all_payments[-1]["payment_due_date"]
        col_13 = all_payments[-1]["payment_at"]
        col_14 = ""
        filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14
        ]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["payment_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE ' + where_expression2

        # Число заявок
        cursor.execute(
            f"""SELECT 
                    COUNT(t0.payment_id)
                FROM public.payments_approval AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT 
                            DISTINCT payment_id,
                            SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                        FROM public.payments_paid_history
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            parent_id::int AS payment_id,
                            parameter_value::numeric AS amount
                        FROM public.payment_draft
                        WHERE page_name = %s AND parameter_name = %s AND user_id = %s
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t0.payment_id = t8.payment_id
                {where_expression2};
                """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-pay', methods=['POST'])
@login_required
def set_paid_payments():
    """Сохранение оплаченных платежей в БД"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Check if the user has access to the "List of contracts" page
        if app_login.current_user.get_role() not in (1, 6):
            return error_handlers.handle403(403)
        if request.method == 'POST':
            selected_rows = request.form.getlist('selectedRows')  # Выбранные столбцы
            contractor_id = request.form.getlist('contractor_id')  # id наших компаний (передаётся id)
            payment_number = request.form.getlist('payment_number')  # Номера платежей (передаётся id)
            payment_pay_sum = request.form.getlist('amount')  # Оплаченные суммы
            payment_full_agreed_status = request.form.getlist('payment_full_agreed_status')  # Сохранить до полной опл.

            selected_rows = [int(i) for i in selected_rows]

            if not sum(selected_rows):
                flash(message=['Не выбрано ни одной заявки', ''], category='error')
                return redirect(url_for('.get_unpaid_payments'))

            contractor_id = [int(i) for i in contractor_id]
            payment_number = [int(i) for i in payment_number]
            payment_pay_sum = [convert_amount(i) for i in payment_pay_sum]
            payment_full_agreed_status = [int(i) for i in payment_full_agreed_status]

            # Данные для удаления временных данных из таблицы payment_draft
            values_p_d = []
            page_name = 'payment-pay'
            parameter_name = 'amount'

            values_b = []  # Список измененных балансов для БД payments_balance
            values_a_u = []  # Список измененных согласованных заявок
            values_a_d = []  # Список удалённых согласованных заявок
            values_p_h = []  # Список оплаченных заявок для записи на БД payments_paid_history
            pay_id_list_raw = []  # Список согласованных id заявок без обработки ошибок
            pay_id_closed = []  # Список закрывающихся заявок

            for i in selected_rows:
                row = i - 1

                if payment_pay_sum[row] is None:
                    flash(message=['Не указана сумма к оплате', f'№ строки {i}'], category='error')
                    return redirect(url_for('.get_unpaid_payments'))

                pay_id_list_raw.append(payment_number[row])

            conn, cursor = app_login.conn_cursor_init_dict()

            # Список балансов компании
            query = """
            SELECT 
                company_id,
                balance_sum
            FROM public.payments_balance;
            """
            cursor.execute(query)
            companies_balance = cursor.fetchall()

            # Список согласованных сумм. Их используем при проверке статуса закрытия платежа
            query = """
            SELECT 
                payment_id, 
                approval_sum
            FROM public.payments_approval 
            WHERE payment_id::int in %s
            """
            execute_values(cursor, query, [pay_id_list_raw])
            approval_sum = cursor.fetchall()

            for i in selected_rows:

                row = i - 1
                status_id = 0

                # Если согласованная сумма больше суммы к оплате и не стоит галка "закрыть после полной оплаты",
                # то статус оплаты - "Частичная оплата с закрытием" (id=11); если галка стоит -"Частичная оплата (id=10)
                # иначе "Полная оплата" (id=9)
                for s in approval_sum:
                    if s[0] == payment_number[row]:
                        if round(float(s[1]) - payment_pay_sum[row], 2) > 0:
                            if i not in payment_full_agreed_status:
                                status_id = 11
                                values_a_d.append((
                                    payment_number[row],
                                ))
                            else:
                                status_id = 10
                                values_a_u.append((
                                    payment_number[row],
                                    float(s[1]) - payment_pay_sum[row]
                                ))
                        else:
                            status_id = 9
                            values_a_d.append(payment_number[row])

                if not status_id:
                    flash(message=['Обновите страницу и повторите попытку снова',
                                   f'Ошибка с платежом {payment_number[row]}'], category='error')
                    app_login.conn_cursor_close(cursor, conn)
                    return redirect(url_for('.get_unpaid_payments'))

                if i not in payment_full_agreed_status:
                    pay_id_closed.append((
                        payment_number[row],
                    ))

                values_p_h.append([
                    payment_number[row],
                    status_id,
                    user_id,
                    payment_pay_sum[row]
                ])

                values_b.append([
                    contractor_id[row],
                    payment_pay_sum[row]
                ])

                values_p_d.append((
                    page_name,
                    payment_number[row],
                    parameter_name,
                    user_id
                ))

            # Пересчитываем баланс компаний
            for val in values_b:
                for com in companies_balance:
                    if val[0] == com[0]:
                        com[1] = float(com[1]) - val[1]

            try:
                # Если есть что записывать в Базу данных
                if values_p_h:
                    # Перезапись в payments_paid_history
                    action_p_h = 'INSERT INTO'
                    table_p_h = 'payments_paid_history'
                    columns_p_h = ('payment_id', 'status_id', 'user_id', 'paid_sum')
                    query_p_h = get_db_dml_query(action=action_p_h, table=table_p_h, columns=columns_p_h)
                    execute_values(cursor, query_p_h, values_p_h)

                    # Обновляем балансы компаний
                    columns_b = ("company_id", "balance_sum")
                    query_b = get_db_dml_query(action='UPDATE', table='payments_balance', columns=columns_b)
                    execute_values(cursor, query_b, companies_balance)

                    # Удаляем временные данные из payment_draft
                    columns_p_d = 'page_name, parent_id::int, parameter_name, user_id::int'
                    query_p_d = get_db_dml_query(action='DELETE', table='payment_draft', columns=columns_p_d)
                    execute_values(cursor, query_p_d, (values_p_d,))

                # Если есть заявки с закрытием
                if values_a_d:
                    columns_a_d = 'payment_id'
                    query_a_d = get_db_dml_query(action='DELETE', table='payments_approval', columns=columns_a_d)
                    execute_values(cursor, query_a_d, (values_a_d,))

                # Если есть заявки с частичным закрытием
                if values_a_u:
                    columns_a_u = ("payment_id", "approval_sum")
                    query_a_u = get_db_dml_query(action='UPDATE', table='payments_approval', columns=columns_a_u)
                    execute_values(cursor, query_a_u, values_a_u)

                flash(message=['Заявки проведены', ''], category='success')

                conn.commit()
                app_login.conn_cursor_close(cursor, conn)

                return redirect(url_for('.get_unpaid_payments'))

            except Exception as e:
                conn.rollback()

                flash(message=['Не указана сумма к оплате', f'№ строки {i}'], category='error')

                app_login.conn_cursor_close(cursor, conn)

                msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')

                return redirect(url_for('.get_unpaid_payments'))

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return redirect(url_for('.get_unpaid_payments'))


@payment_app_bp.route('/payment-approval-list', methods=['GET'])
@login_required
def get_payments_approval_list():
    """Выгрузка из БД списка оплаченных платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, log_description=request.method, user_id=user_id,
                               ip_address=app_login.get_client_ip())
        role = app_login.current_user.get_role()

        # Check if the user has access to the "List of contracts" page
        if role not in (1, 4, 6):
            return error_handlers.handle403(403)

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        cursor.execute(
            """
            SELECT 
                t0.payment_id + 1 AS payment_id,
                (t8.created_at::timestamp without time zone + interval '1 day')::text AS created_at
            FROM public.payments_approval AS t0
            LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            created_at
                        FROM public.payments_approval_history
                        ORDER BY payment_id, created_at DESC
                ) AS t8 ON t0.payment_id = t8.payment_id
            ORDER BY t8.created_at DESC, t0.payment_id DESC
            LIMIT 1;
            """
        )
        all_payments = cursor.fetchall()

        # Согласован и не оплачено
        cursor.execute(
            """WITH
                t2 AS (SELECT
                        COALESCE(sum(approval_sum), 0) AS approval_sum
                    FROM public.payments_approval)
                SELECT
                    t2.approval_sum AS approval_money,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t2.approval_sum, 0), '999 999 990D99 ₽')) AS approval_money_rub
                FROM t2;"""
        )
        money = cursor.fetchone()

        # Список ответственных
        cursor.execute(
            "SELECT user_id, last_name, first_name FROM public.users WHERE is_fired = FALSE ORDER BY last_name, first_name")
        responsible = cursor.fetchall()

        # Список типов заявок
        cursor.execute(
            """SELECT
                cost_item_id,
                cost_item_name,
                cost_item_category
            FROM public.payment_cost_items
            ORDER BY cost_item_category, cost_item_name""")
        cost_items_list = cursor.fetchall()
        # передаём данные в виде словаря для создания сгруппированного выпадающего списка
        cost_items = {}
        for item in cost_items_list:
            key = item[2]
            value = [item[1], item[0]]
            if key in cost_items:
                cost_items[key].append(value)
            else:
                cost_items[key] = [value]

        # Список объектов
        cursor.execute("SELECT object_id, object_name FROM public.objects ORDER BY object_name")
        objects_name = cursor.fetchall()

        # Список контрагентов
        cursor.execute("SELECT DISTINCT partner FROM public.payments_summary_tab ORDER BY partner")
        partners = cursor.fetchall()

        # Список наших компаний из таблицы contractors
        cursor.execute("SELECT contractor_id, contractor_name FROM public.our_companies")
        our_companies = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [12, 1, all_payments[-1]['created_at']],  # Первая колонка
                'col_id': all_payments[-1]['payment_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        tab_rows = 1

        # Список разделов в меню платежей
        payment_approval_menu = get_payment_approval_menu(role)
        if role in (1, 4, 6):
            payment_approval_menu[1]['class'] = 'focus_button'

        return render_template('payment-approval-list.html', menu=hlink_menu,
                               menu_profile=hlink_profile, sort_col=sort_col, responsible=responsible,
                               cost_items=cost_items, objects_name=objects_name, partners=partners,
                               our_companies=our_companies, money=money, tab_rows=tab_rows, setting_users=setting_users,
                               nonce=get_nonce(), payment_approval_menu=payment_approval_menu,
                               title='Согласованные платежи')

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-paymentApprovalList-pagination', methods=['POST'])
@login_required
def get_payment_approval_list_pagination():
    """Постраничная выгрузка списка согласованных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-approval-list'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't0.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()
        # print(f"""
        #                     WHERE {where_expression}
        #                     ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
        # """)
        # print(query_value)
        try:
            cursor.execute(
                f"""SELECT 
                        t0.payment_id,
                        t1.payment_number,
                        t3.contractor_name,
                        t3.contractor_id,
                        t4.cost_item_name,
                        SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,
                        t1.basis_of_payment,  
                        t5.first_name,
                        t5.last_name,
                        SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                        t1.payment_description, 
                        COALESCE(t6.object_name, '') AS object_name,
                        t1.partner,
                        t1.payment_sum, 
                        TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                        t0.approval_sum,
                        TRIM(BOTH ' ' FROM to_char(t0.approval_sum, '999 999 990D99 ₽')) AS approval_sum_rub,
                        COALESCE(t7.paid_sum, '0') AS paid_sum, 
                        TRIM(BOTH ' ' FROM to_char(COALESCE(t7.paid_sum, 0), '999 999 990D99 ₽')) AS paid_sum_rub,
                        to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                        t1.payment_due_date::text AS payment_due_date,
                        to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt,
                        t1.payment_at::timestamp without time zone::text AS payment_at,
                        to_char(t8.created_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS created_at_txt,
                        t8.created_at::timestamp without time zone::text AS created_at
                    FROM public.payments_approval AS t0
                    LEFT JOIN (
                        SELECT 
                            payment_id, 
                            payment_number, 
                            basis_of_payment,
                            payment_description,
                            partner,
                            payment_sum,
                            payment_due_date,
                            payment_at,
                            our_companies_id,
                            cost_item_id,
                            responsible,
                            object_id
                        FROM public.payments_summary_tab
                    ) AS t1 ON t0.payment_id = t1.payment_id
                    LEFT JOIN (
                        SELECT contractor_id,
                            contractor_name
                        FROM public.our_companies            
                    ) AS t3 ON t1.our_companies_id = t3.contractor_id
                    LEFT JOIN (
                        SELECT cost_item_id,
                            cost_item_name
                        FROM public.payment_cost_items            
                    ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                    LEFT JOIN (
                            SELECT user_id,
                                first_name,
                                last_name
                            FROM public.users
                    ) AS t5 ON t1.responsible = t5.user_id
                    LEFT JOIN (
                            SELECT object_id,
                                object_name
                            FROM public.objects
                    ) AS t6 ON t1.object_id = t6.object_id
                    LEFT JOIN (
                            SELECT 
                                DISTINCT payment_id,
                                SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                            FROM public.payments_paid_history
                    ) AS t7 ON t0.payment_id = t7.payment_id
                    LEFT JOIN (
                                SELECT DISTINCT ON (payment_id) 
                                    payment_id,
                                    created_at
                                FROM public.payments_approval_history
                                ORDER BY payment_id, created_at DESC
                        ) AS t8 ON t0.payment_id = t8.payment_id
                    WHERE {where_expression}
                    ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                    LIMIT {limit};
                """,
                query_value
            )
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = all_payments[-1]["payment_number"]
        col_1 = all_payments[-1]["cost_item_name"]
        col_2 = all_payments[-1]["basis_of_payment"]
        col_3 = f'{all_payments[-1]["contractor_name"]}: {all_payments[-1]["payment_description"]}'
        col_4 = all_payments[-1]["object_name"]
        col_5 = f'{all_payments[-1]["last_name"]} {all_payments[-1]["first_name"]}'
        col_6 = all_payments[-1]["partner"]
        col_7 = all_payments[-1]["payment_sum"]
        col_8 = all_payments[-1]["approval_sum"]
        col_9 = all_payments[-1]["paid_sum"]
        col_10 = all_payments[-1]["payment_due_date"]
        col_11 = all_payments[-1]["payment_at"]
        col_12 = all_payments[-1]["created_at"]
        filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12
        ]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["payment_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE ' + where_expression2

        # Число заявок
        cursor.execute(
            f"""SELECT 
                    COUNT(t0.payment_id)
                FROM public.payments_approval AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT 
                            DISTINCT payment_id,
                            SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                        FROM public.payments_paid_history
                ) AS t7 ON t0.payment_id = t7.payment_id
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            created_at
                        FROM public.payments_approval_history
                ) AS t8 ON t0.payment_id = t8.payment_id
                {where_expression2};
            """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-paid-list', methods=['GET'])
@login_required
def get_payments_paid_list():
    """Выгрузка из БД списка оплаченных платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())
        role = app_login.current_user.get_role()

        # Check if the user has access to the "List of contracts" page
        if role not in (1, 4, 6):
            return error_handlers.handle403(403)

        # # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        cursor.execute(
            """
            WITH t0 AS (
                SELECT 
                    payment_id
                FROM public.payments_paid_history
                GROUP BY payment_id
            )
            SELECT 
                t0.payment_id + 1 AS payment_id,
                (t1.payment_at::timestamp without time zone + interval '1 day')::text AS payment_at
            FROM t0
            LEFT JOIN (
                        SELECT 
                            payment_id,
                            payment_at
                        FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
            ORDER BY t1.payment_at DESC, t0.payment_id DESC
            LIMIT 1;
            """
        )
        all_payments = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [12, 1, all_payments[-1]['payment_at']],  # Первая колонка - DESC
                'col_id': all_payments[-1]['payment_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        tab_rows = 1

        # Список разделов в меню платежей
        payment_approval_menu = get_payment_approval_menu(role)
        if role in (1, 4, 6):
            payment_approval_menu[2]['class'] = 'focus_button'

        return render_template('payment-paid-list.html', menu=hlink_menu, menu_profile=hlink_profile,
                               sort_col=sort_col, tab_rows=tab_rows, setting_users=setting_users, nonce=get_nonce(),
                               payment_approval_menu=payment_approval_menu, title='Оплаченные платежи')

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-paymentPaidList-pagination', methods=['POST'])
@login_required
def get_payment_paid_list_pagination():
    """Постраничная выгрузка списка оплаченных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-paid-list'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't1.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            cursor.execute(
                f"""WITH t0 AS (
                        SELECT 
                            payment_id,
                            MAX(created_at) AS paid_at,
                            SUM(paid_sum) AS paid_sum
                        FROM public.payments_paid_history
                        GROUP BY payment_id
                    )
                    SELECT 
                            t0.payment_id,
                            t1.payment_number,
                            t3.contractor_name,
                            t4.cost_item_name,
                            SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,
                            t1.basis_of_payment,  
                            t5.first_name,
                            t5.last_name,
                            SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                            t1.payment_description,  
                            COALESCE(t6.object_name, '') AS object_name,
                            t1.partner,
                            t1.payment_sum,
                            TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                            COALESCE(t2.approval_sum, '0') AS approval_sum,
                            TRIM(BOTH ' ' FROM to_char(t2.approval_sum, '999 999 990D99 ₽')) AS approval_sum_rub,
                            t0.paid_sum AS paid_sum,
                            TRIM(BOTH ' ' FROM to_char(COALESCE(t0.paid_sum, 0), '999 999 990D99 ₽')) AS paid_sum_rub,
                            to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                            t1.payment_due_date::text AS payment_due_date,
                            t8.status_name,
                            to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt, 
                            t1.payment_at::timestamp without time zone::text AS payment_at
                        FROM t0
                        LEFT JOIN (
                            SELECT 
                                payment_id, 
                                payment_number, 
                                basis_of_payment,
                                payment_description,
                                partner,
                                payment_sum,
                                payment_due_date,
                                payment_at,
                                our_companies_id,
                                cost_item_id,
                                responsible,
                                object_id
                            FROM public.payments_summary_tab
                        ) AS t1 ON t0.payment_id = t1.payment_id
                        LEFT JOIN (
                                SELECT DISTINCT ON (payment_id) 
                                    payment_id,
                                    SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                                FROM public.payments_approval_history
                                ORDER BY payment_id, created_at DESC
                        ) AS t2 ON t0.payment_id = t2.payment_id
                        LEFT JOIN (
                            SELECT contractor_id,
                                contractor_name
                            FROM public.our_companies            
                        ) AS t3 ON t1.our_companies_id = t3.contractor_id
                        LEFT JOIN (
                            SELECT cost_item_id,
                                cost_item_name
                            FROM public.payment_cost_items            
                        ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                        LEFT JOIN (
                                SELECT user_id,
                                    first_name,
                                    last_name
                                FROM public.users
                        ) AS t5 ON t1.responsible = t5.user_id
                        LEFT JOIN (
                                SELECT object_id,
                                    object_name
                                FROM public.objects
                        ) AS t6 ON t1.object_id = t6.object_id
                        LEFT JOIN (
                                SELECT DISTINCT ON (payment_id) 
                                    payment_id,
                                    status_id
                                FROM public.payments_paid_history
                                ORDER BY payment_id, created_at DESC
                        ) AS t7 ON t0.payment_id = t7.payment_id
                        LEFT JOIN (
                                SELECT payment_agreed_status_id AS status_id,
                                    payment_agreed_status_name AS status_name
                                FROM public.payment_agreed_statuses
                        ) AS t8 ON t7.status_id = t8.status_id
                    WHERE {where_expression}
                    ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                    LIMIT {limit};
                """,
                query_value
            )
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = ""
        col_1 = all_payments[-1]["payment_number"]
        col_2 = all_payments[-1]["cost_item_name"]
        col_3 = all_payments[-1]["basis_of_payment"]
        col_4 = f'{all_payments[-1]["contractor_name"]}: {all_payments[-1]["payment_description"]}'
        col_5 = all_payments[-1]["object_name"]
        col_6 = f'{all_payments[-1]["last_name"]} {all_payments[-1]["first_name"]}'
        col_7 = all_payments[-1]["partner"]
        col_8 = all_payments[-1]["payment_sum"]
        col_9 = all_payments[-1]["approval_sum"]
        col_10 = all_payments[-1]["paid_sum"]
        col_11 = all_payments[-1]["payment_due_date"]
        col_12 = all_payments[-1]["payment_at"]
        col_13 = all_payments[-1]["status_name"]
        filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["payment_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE ' + where_expression2

        # Число заявок
        cursor.execute(
            f"""WITH t0 AS (
                    SELECT 
                        payment_id,
                        MAX(created_at) AS paid_at,
                        SUM(paid_sum) AS paid_sum
                    FROM public.payments_paid_history
                    GROUP BY payment_id)
                SELECT 
                        COUNT(t0.payment_id)
                    FROM t0
                    LEFT JOIN (
                        SELECT 
                            payment_id, 
                            payment_number, 
                            basis_of_payment,
                            payment_description,
                            partner,
                            payment_sum,
                            payment_due_date,
                            payment_at,
                            our_companies_id,
                            cost_item_id,
                            responsible,
                            object_id
                        FROM public.payments_summary_tab
                    ) AS t1 ON t0.payment_id = t1.payment_id
                    LEFT JOIN (
                            SELECT DISTINCT ON (payment_id) 
                                payment_id,
                                SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                            FROM public.payments_approval_history
                            ORDER BY payment_id, created_at DESC
                    ) AS t2 ON t0.payment_id = t2.payment_id
                    LEFT JOIN (
                        SELECT contractor_id,
                            contractor_name
                        FROM public.our_companies            
                    ) AS t3 ON t1.our_companies_id = t3.contractor_id
                    LEFT JOIN (
                        SELECT cost_item_id,
                            cost_item_name
                        FROM public.payment_cost_items            
                    ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                    LEFT JOIN (
                            SELECT user_id,
                                first_name,
                                last_name
                            FROM public.users
                    ) AS t5 ON t1.responsible = t5.user_id
                    LEFT JOIN (
                            SELECT object_id,
                                object_name
                            FROM public.objects
                    ) AS t6 ON t1.object_id = t6.object_id
                    LEFT JOIN (
                            SELECT DISTINCT ON (payment_id) 
                                payment_id,
                                status_id
                            FROM public.payments_paid_history
                            ORDER BY payment_id, created_at DESC
                    ) AS t7 ON t0.payment_id = t7.payment_id
                    LEFT JOIN (
                            SELECT payment_agreed_status_id AS status_id,
                                payment_agreed_status_name AS status_name
                            FROM public.payment_agreed_statuses
                    ) AS t8 ON t7.status_id = t8.status_id
            {where_expression2};
            """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-list', methods=['GET'])
@login_required
def get_payments_list():
    """Выгрузка из БД списка оплаченных платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        cursor.execute(
            """
                SELECT 
                    t1.payment_id + 1 AS payment_id,
                    (t1.payment_at::timestamp without time zone + interval '1 day')::text AS payment_at
                FROM public.payments_summary_tab AS t1
                WHERE t1.payment_owner = %s OR t1.responsible = %s
                ORDER BY t1.payment_at DESC, t1.payment_id DESC
                LIMIT 1;
                """,
            [user_id, user_id]
        )
        all_payments = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [10, 1, all_payments[-1]['payment_at']],  # Первая колонка - ASC
                'col_id': all_payments[-1]['payment_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        tab_rows = 1

        return render_template('payment-list.html', menu=hlink_menu, menu_profile=hlink_profile, sort_col=sort_col,
                               tab_rows=tab_rows, setting_users=setting_users, nonce=get_nonce(),
                               title='Список платежей')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-paymentList-pagination', methods=['POST'])
@login_required
def get_payment_list_pagination():
    """Постраничная выгрузка списка созданных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-list'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't1.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })
        # print('/get-contractList-pagination\n', '= - ' * 20,
        #       f"""WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND {where_expression}
        #         ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
        #         LIMIT {limit};""")
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            cursor.execute(
                f"""SELECT 
                        t1.payment_id,
                        t3.contractor_name, 
                        t4.cost_item_name, 
                        t1.payment_number,  
                        SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short, 
                        t1.basis_of_payment, 
                        t5.first_name,
                        t5.last_name,
                        SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                        t1.payment_description, 
                        COALESCE(t6.object_name, '') AS object_name,
                        t1.partner,
                        t1.payment_sum,
                        TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                        to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                        t1.payment_due_date::text AS payment_due_date,
                        to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt,
                        t1.payment_at::timestamp without time zone::text AS payment_at,
                        COALESCE(t7.paid_sum, 0) AS paid_sum,
                        TRIM(BOTH ' ' FROM to_char(COALESCE(t7.paid_sum, 0), '999 999 990D99 ₽')) AS paid_sum_rub,
                        t2.available_for_deletion
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT payment_id,
                            COUNT(*) = 1 AS available_for_deletion
                        FROM public.payments_approval_history
                        GROUP BY payment_id
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                            SELECT 
                                DISTINCT payment_id,
                                SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                            FROM public.payments_paid_history
                    ) AS t7 ON t1.payment_id = t7.payment_id
                WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND {where_expression}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = all_payments[-1]["payment_number"]
        col_1 = all_payments[-1]["cost_item_name"]
        col_2 = all_payments[-1]["basis_of_payment"]
        col_3 = f'{all_payments[-1]["contractor_name"]}: {all_payments[-1]["payment_description"]}'
        col_4 = all_payments[-1]["object_name"]
        col_5 = f'{all_payments[-1]["last_name"]} {all_payments[-1]["first_name"]}'
        col_6 = all_payments[-1]["partner"]
        col_7 = all_payments[-1]["payment_sum"]
        col_8 = all_payments[-1]["paid_sum"]
        col_9 = all_payments[-1]["payment_due_date"]
        col_10 = all_payments[-1]["payment_at"]
        filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10
        ]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["payment_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE (t1.payment_owner = %s OR t1.responsible = %s) AND ' + where_expression2
        else:
            where_expression2 = 'WHERE (t1.payment_owner = %s OR t1.responsible = %s)'

        # Число заявок
        cursor.execute(
            f"""SELECT 
                    COUNT(t1.payment_id)
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                        SELECT DISTINCT ON (payment_id) 
                            payment_id,
                            status_id,
                            SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                        FROM public.payments_approval_history
                        ORDER BY payment_id, created_at DESC
                ) AS t2 ON t1.payment_id = t2.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                            SELECT 
                                DISTINCT payment_id,
                                SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                            FROM public.payments_paid_history
                    ) AS t7 ON t1.payment_id = t7.payment_id
                {where_expression2};
                """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-inflow-history-list', methods=['GET'])
@login_required
def get_inflow_history_list():
    """Выгрузка из БД списка входящих платежей"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())
        role = app_login.current_user.get_role()

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        cursor.execute(
            """
                SELECT 
                    t1.inflow_id + 1 AS inflow_id,
                    (t1.inflow_at::timestamp without time zone + interval '1 day')::text AS inflow_at
                FROM public.payments_inflow_history AS t1
                ORDER BY t1.inflow_at DESC, t1.inflow_at DESC
                LIMIT 1;
                """
        )
        all_payments = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Список колонок для сортировки
        if len(all_payments):
            sort_col = {
                'col_1': [5, 1, all_payments[-1]['inflow_at']],  # Первая колонка - DESC
                'col_id': all_payments[-1]['inflow_id']
            }
        else:
            sort_col = {
                'col_1': [False, 1, False],  # Первая колонка
                'col_id': False
            }

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        tab_rows = 1

        # Список разделов в меню платежей
        payment_approval_menu = get_payment_approval_menu(role)
        if role in (1, 4, 6):
            payment_approval_menu[3]['class'] = 'focus_button'

        return render_template('payment-inflow-history-list.html', menu=hlink_menu,
                               menu_profile=hlink_profile, sort_col=sort_col, tab_rows=tab_rows,
                               setting_users=setting_users, nonce=get_nonce(),
                               payment_approval_menu=payment_approval_menu, title='История входящих платежей')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-inflowHistory-pagination', methods=['POST'])
@login_required
def get_inflow_history_list_pagination():
    """Постраничная выгрузка списка входящих платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-inflow-history-list'
        limit = request.get_json()['limit']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't1.inflow_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if col_1.split('#')[0] == 'False':
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': 'Нет данных',
            })

        sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
            query_value, sort_col, col_num = \
            get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id)

        # Когда происходит горизонтальный скролл страницы и нажимается кнопка сортировки, вызывается
        # дополнительная пагинация с пустыми значениями сортировки. Отлавливаем этот случай, ничего не делаем
        if not col_1_val and not col_id_val:
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Skip pagination with empty sort data',
            })
        # print('/get-inflowHistory-pagination\n', '= - ' * 20, '\n',
        #       f"""WHERE {where_expression}
        #         ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
        #         LIMIT {limit};""")
        # print('query_value', query_value)
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            cursor.execute(
                f"""SELECT 
                        t1.inflow_id,
                        t2.contractor_name, 
                        t1.inflow_description, 
                        t3.inflow_type_name, 
                        t1.inflow_sum, 
                        TRIM(BOTH ' ' FROM to_char(t1.inflow_sum, '999 999 990D99 ₽')) AS inflow_sum_rub,
                        to_char(t1.inflow_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS inflow_at_txt,
                        t1.inflow_at::timestamp without time zone::text AS inflow_at
                FROM public.payments_inflow_history AS t1
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t2 ON t1.inflow_company_id = t2.contractor_id
                LEFT JOIN (
                    SELECT inflow_type_id,
                        inflow_type_name
                    FROM public.payment_inflow_type
                ) AS t3 ON t1.inflow_type_id = t3.inflow_type_id
                WHERE {where_expression}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order}
                LIMIT {limit};
                """,
                query_value
            )
            # pprint(cursor.query)
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': msg_for_user,
            })

        if not len(all_payments):
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': sort_col,
                'status': 'success',
                'description': 'Конец таблицы. Ничего не найдено',
            })

        col_0 = all_payments[-1]["inflow_id"]
        col_1 = all_payments[-1]["contractor_name"]
        col_2 = all_payments[-1]["inflow_description"]
        col_3 = all_payments[-1]["inflow_type_name"]
        col_4 = all_payments[-1]["inflow_sum"]
        col_5 = all_payments[-1]["inflow_at"]
        filter_col = [col_0, col_1, col_2, col_3, col_4, col_5]

        # Список колонок для сортировки, добавляем последние значения в столбах сортировки
        sort_col['col_1'].append(filter_col[col_num])
        sort_col['col_id'] = all_payments[-1]["inflow_id"]

        for i in range(len(all_payments)):
            all_payments[i] = dict(all_payments[i])

        if where_expression2:
            where_expression2 = 'WHERE ' + where_expression2

        # Число заявок
        cursor.execute(
            f"""SELECT 
                    COUNT(t1.inflow_id)
                FROM public.payments_inflow_history AS t1
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t2 ON t1.inflow_company_id = t2.contractor_id
                LEFT JOIN (
                    SELECT inflow_type_id,
                        inflow_type_name
                    FROM public.payment_inflow_type
                ) AS t3 ON t1.inflow_type_id = t3.inflow_type_id
                {where_expression2};
                """,
            query_value
        )
        tab_rows = cursor.fetchone()[0]

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': tab_rows,
            'page': page_name,
            'setting_users': setting_users,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-list-export_to_excel')
@login_required
def payment_list_export_to_excel():
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            # Fetch data from db
            cursor.execute(
                f"""SELECT 
                        t1.payment_number,
                        t4.cost_item_name, 
                        SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,
                        CONCAT(
                            t3.contractor_name, ': ', SUBSTRING(t1.payment_description, 1,70)
                            ) AS payment_description_short,
                        COALESCE(t6.object_name, '') AS object_name,
                        CONCAT_WS(' ', t5.last_name, t5.first_name) AS responsible,
                        t1.partner,
                        t1.payment_sum,
                        COALESCE(t7.paid_sum, 0) AS paid_sum,
                        to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date,
                        to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at
                FROM public.payments_summary_tab AS t1
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                            SELECT 
                                DISTINCT payment_id,
                                SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                            FROM public.payments_paid_history
                    ) AS t7 ON t1.payment_id = t7.payment_id
                WHERE (t1.payment_owner = %s OR t1.responsible = %s);
                """,
                [user_id, user_id]
            )
            all_payments = cursor.fetchall()

            app_login.conn_cursor_close(cursor, conn)

            # Create a Pandas DataFrame from the query result
            columns = [
                '№ платежа',
                'Статья затрат',
                'Наименование платежа',
                'Описание',
                'Объект',
                'Ответственный',
                'Контрагент',
                'Общая сумма',
                'Оплаченная сумма',
                'Срок оплаты',
                'Дата создания',
            ]
            df = pd.DataFrame(all_payments, columns=columns)

            # Create Excel workbook and add data
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_name = f"Список_платежей - {timestamp}.xlsx"

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=file_name) as temp_file:
                excel_file_path = temp_file.name

                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active

                # Write the column headers to the Excel file
                for col_num, value in enumerate(columns, 1):
                    ws.cell(row=1, column=col_num, value=value)

                # Write the data to the Excel file
                for row_num, row_data in enumerate(all_payments, 2):
                    for col_num, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)

                wb.save(excel_file_path)

            wb.close()

            return send_file(excel_file_path, as_attachment=True, download_name=file_name)

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            return jsonify({
                'status': 'error',
                'description': msg_for_user,
            })

    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/get_card_payment/<page_url>/<int:payment_id>', methods=['GET'])
@login_required
def get_card_payment(page_url, payment_id):
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, log_description=payment_id,
                               user_id=user_id, ip_address=app_login.get_client_ip())

        data = payment_id
        page_url = page_url

        user_role = app_login.current_user.get_role()
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()
        # payment_id = data['paymentId']
        payment_id = data

        # Список ответственных
        cursor.execute(
            "SELECT user_id, last_name, first_name FROM public.users WHERE is_fired = FALSE")
        responsible = cursor.fetchall()

        # Список типов заявок
        cursor.execute(
            """SELECT 
                        cost_item_id, 
                        cost_item_name, 
                        cost_item_category 
                    FROM public.payment_cost_items 
                    ORDER BY cost_item_category, cost_item_name""")
        cost_items_list = cursor.fetchall()

        # передаём данные в виде словаря для создания сгруппированного выпадающего списка
        cost_items = {}
        for item in cost_items_list:
            key = item[2]
            value = [item[1], item[0]]
            if key in cost_items:
                cost_items[key].append(value)
            else:
                cost_items[key] = [value]

        # Список объектов
        cursor.execute("SELECT object_id, object_name FROM public.objects")
        objects_name = cursor.fetchall()

        # Список контрагентов
        cursor.execute("SELECT DISTINCT partner FROM public.payments_summary_tab")
        partners = cursor.fetchall()

        # Список наших компаний из таблицы contractors
        cursor.execute("SELECT contractor_id, contractor_name FROM public.our_companies")
        our_companies = cursor.fetchall()

        # Все поля из формы по заявке из карточки
        cursor.execute(
            """SELECT 
                    t1.payment_id,
                    t3.contractor_name, 
                    t3.contractor_id, 
                    t4.cost_item_id,
                    t4.cost_item_name, 
                    t1.payment_number,  
                    SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,
                    t1.basis_of_payment, 
                    t5.user_id,
                    t5.first_name,
                    t5.last_name,
                    SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                    t1.payment_description,
                    COALESCE(t6.object_id, 0) AS object_id,
                    COALESCE(t6.object_name, '') AS object_name,
                    t1.partner,
                    t1.payment_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                    COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum) AS unapproved_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum), '999 999 990D99 ₽')) AS unapproved_sum_rub,
                    COALESCE(t2.approval_sum, 0) AS unpaid_approval_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t2.approval_sum, 0), '999 999 990D99 ₽')) AS unpaid_approval_sum_rub,
                    COALESCE(t9.approval_sum, 0) AS approval_to_pay_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t9.approval_sum, 0), '999 999 990D99 ₽')) AS approval_to_pay_sum_rub,
                    CAST(t1.payment_due_date AS TEXT) AS payment_due_date,
                    t2.status_id,
                    date_trunc('second', t1.payment_at::timestamp without time zone)::text AS payment_at,
                    date_trunc('second', t1.payment_at::timestamp without time zone) AS payment_at2,
                    t1.payment_full_agreed_status
            FROM public.payments_summary_tab AS t1
            LEFT JOIN (
                    SELECT DISTINCT ON (payment_id) 
                        payment_id,
                        status_id,
                        SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                    FROM public.payments_approval_history
                    ORDER BY payment_id, created_at DESC
            ) AS t2 ON t1.payment_id = t2.payment_id
            LEFT JOIN (
                SELECT contractor_id,
                    contractor_name
                FROM public.our_companies            
            ) AS t3 ON t1.our_companies_id = t3.contractor_id
            LEFT JOIN (
                SELECT cost_item_id,
                    cost_item_name
                FROM public.payment_cost_items            
            ) AS t4 ON t1.cost_item_id = t4.cost_item_id
            LEFT JOIN (
                    SELECT user_id,
                        first_name,
                        last_name
                    FROM public.users
            ) AS t5 ON t1.responsible = t5.user_id
            LEFT JOIN (
                    SELECT object_id,
                        object_name
                    FROM public.objects
            ) AS t6 ON t1.object_id = t6.object_id
            LEFT JOIN (
                    SELECT 
                        payment_id,
                        approval_sum
                    FROM public.payments_approval
            ) AS t9 ON t1.payment_id = t9.payment_id
            WHERE t1.payment_id = %s
            ORDER BY t1.payment_due_date;
            """,
            [payment_id]
        )
        payment = cursor.fetchone()

        # Список согласованных платежей
        cursor.execute(
            """WITH
                t0 AS (SELECT 
                    payment_id,
                    SUM(approval_sum) AS approval_sum
                          FROM public.payments_approval_history
                          GROUP BY payment_id)
            SELECT 
                    t1.payment_id,
                    date_trunc('second', t1.created_at::timestamp without time zone)::text AS payment_at,
                    t2.payment_agreed_status_name, 
                    t0.approval_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.approval_sum, '999 999 990D99 ₽')) AS approval_sum_rub
            FROM public.payments_approval_history AS t1
            LEFT JOIN (
                    SELECT  
                        payment_agreed_status_id,
                        payment_agreed_status_name
                    FROM public.payment_agreed_statuses
            ) AS t2 ON t1.status_id = t2.payment_agreed_status_id
            LEFT JOIN t0 ON t1.payment_id = t0.payment_id
    
            WHERE t1.payment_id = %s
            ORDER BY t1.created_at;
            """,
            [payment_id]
        )
        approval = cursor.fetchall()

        # Список оплаченных платежей
        cursor.execute(
            """WITH
                t0 AS (SELECT 
                    payment_id,
                    SUM(paid_sum) AS paid_sum
                          FROM public.payments_paid_history
                          GROUP BY payment_id)
            SELECT 
                    t1.payment_id,
                    to_char(t1.created_at, 'dd.MM.yy HH24:MI:SS') AS payment_at_2,
                    date_trunc('second', t1.created_at::timestamp without time zone)::text AS payment_at,
                    t2.payment_agreed_status_name, 
                    t0.paid_sum AS total_paid_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.paid_sum, '999 999 990D99 ₽')) AS paid_sum_rub,
                    TRIM(BOTH ' ' FROM to_char(t0.paid_sum, '999 999 990D99 ₽')) AS total_paid_sum_rub
            FROM public.payments_paid_history AS t1
            LEFT JOIN (
                    SELECT  
                        payment_agreed_status_id,
                        payment_agreed_status_name
                    FROM public.payment_agreed_statuses
            ) AS t2 ON t1.status_id = t2.payment_agreed_status_id
            LEFT JOIN t0 ON t1.payment_id = t0.payment_id
    
            WHERE t1.payment_id = %s
            ORDER BY t1.created_at DESC;
            """,
            [payment_id]
        )
        paid = cursor.fetchall()

        # Лог платежа
        cursor.execute(
            """
            WITH 
            t1 AS (
                SELECT 
                    created_at,
                    'Согласование' AS type,
                    status_id,
                    TRIM(BOTH ' ' FROM to_char(approval_sum::numeric, '999 999 990D99 ₽')) AS sum
                FROM public.payments_approval_history
                WHERE payment_id = %s
                UNION ALL 
                    SELECT  
                        created_at,
                        'Оплата' AS type,
                        status_id,
                        TRIM(BOTH ' ' FROM to_char(paid_sum::numeric, '999 999 990D99 ₽')) AS sum
                    FROM public.payments_paid_history
                    WHERE payment_id = %s),
            t2 AS (
                    SELECT  
                        payment_agreed_status_id,
                        payment_agreed_status_name
                    FROM public.payment_agreed_statuses
            )
            SELECT 
                to_char(t1.created_at, 'dd.MM.yy') AS created_at_date,
                to_char(t1.created_at, 'HH24:MI:SS') AS created_at_time,
                t1.type,
                t2.payment_agreed_status_name,
                COALESCE(t1.sum, '') AS sum
            FROM t1
            LEFT JOIN t2 ON t1.status_id=t2.payment_agreed_status_id
            ORDER BY t1.created_at
                    ;
            """,
            [payment_id, payment_id]
        )
        logs = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        # Return the updated data as a response
        return jsonify({
            'status': 'success',
            'payment': dict(payment),
            'approval': approval,
            'paid': paid,
            'responsible': responsible,
            'cost_items': cost_items,
            'objects_name': objects_name,
            'partners': partners,
            'our_companies': our_companies,
            'logs': logs,
            'user_role': user_role
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/save_payment', methods=['POST'])
@login_required
def save_payment():
    """Сохраняем изменения платежа из карточки платежа"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_url = request.get_json()['page_url']  # страница с которой запущено сохранение
        payment_id = int(request.get_json()['payment_id'])  # Номера платежей (передаётся id)

        basis_of_payment = request.get_json()['basis_of_payment']  # Основание (наименование) платежа
        responsible = int(request.get_json()['responsible'])  # id ответственного
        cost_item_id = int(request.get_json()['cost_item_id'])  # id статьи затрат
        object_id = int(request.get_json()['object_id']) if request.get_json()['object_id'] else 0  # id объекта
        payment_description = request.get_json()['payment_description']  # Описание
        partners = request.get_json()['partners']  # Контрагент
        payment_due_date = date.fromisoformat(request.get_json()['payment_due_date'])  # Срок оплаты
        our_companies_id = int(request.get_json()['our_company_id'])  # id нашей компании
        payment_sum = convert_amount(request.get_json()['payment_sum'])  # Общая сумма
        approval_sum = convert_amount(request.get_json()['sum_approval'])  # Сумма согласования
        payment_full_agreed_status = request.get_json()['payment_full_agreed_status']  # Сохранить до полной оплаты

        # data-value данные для проверки, в каких полях были изменения
        basis_of_payment_dataset = request.get_json()['basis_of_payment_dataset']  # Основание (наименование) платежа
        responsible_dataset = int(request.get_json()['responsible_dataset'])  # id ответственного
        cost_item_id_dataset = int(request.get_json()['cost_item_id_dataset'])  # id статьи затрат
        object_id_dataset = int(request.get_json()['object_id_dataset'])  # id объекта
        payment_description_dataset = request.get_json()['payment_description_dataset']  # Описание
        partners_dataset = request.get_json()['partners_dataset']  # Контрагент
        payment_due_date_dataset = date.fromisoformat(request.get_json()['payment_due_date_dataset'])  # Срок оплаты
        our_companies_id_dataset = int(request.get_json()['our_company_id_dataset'])  # id нашей компании
        payment_sum_dataset = convert_amount(request.get_json()['payment_sum_dataset'])  # Общая сумма
        approval_sum_dataset = convert_amount(request.get_json()['sum_approval_dataset'])  # Сумма согласования
        payment_full_agreed_status_dataset = request.get_json()['p_full_agreed_s_dataset']  # Сохранить до полной оплаты
        payment_full_agreed_status_dataset = True if payment_full_agreed_status_dataset == 'true' else False

        # На листе оплата платежей часть данных изменить нельзя
        if page_url == 'payment-pay':
            cost_item_id = cost_item_id_dataset
            payment_sum = payment_sum_dataset
            approval_sum = approval_sum_dataset
            payment_full_agreed_status = payment_full_agreed_status_dataset
            object_id = object_id_dataset

        status_id = 12  # Статус заявки ("Обновлено")
        approval_sum_p_a_h = 0  # Сумма согласования
        values_p_s_t = []  # Данные для записи в таблицу payments_summary_tab
        values_p_a = []  # Данные для записи в таблицу payments_approval
        values_p_a_h = []  # Данные для записи в таблицу payments_approval_history
        columns_p_s_t = []  # Список колонок, в которых произошло изменение (payments_summary_tab)
        columns_p_a = []  # Список колонок, в которых произошло изменение (payments_approval)
        columns_p_a_h = []  # Список колонок, в которых произошло изменение (payments_approval_history)

        # Словарь True/False для определения, что пользователь изменил в payments_summary_tab
        status_change_list_p_s_t = {}
        status_change_list_p_s_t['basis_of_payment'] = [
            basis_of_payment != basis_of_payment_dataset,
            basis_of_payment
        ]
        status_change_list_p_s_t['responsible'] = [
            responsible != responsible_dataset,
            responsible
        ]
        status_change_list_p_s_t['cost_item_id'] = [
            cost_item_id != cost_item_id_dataset,
            cost_item_id
        ]
        status_change_list_p_s_t['object_id::smallint'] = [
            object_id != object_id_dataset,
            None if not object_id else object_id
        ]
        status_change_list_p_s_t['payment_description'] = [
            payment_description != payment_description_dataset,
            payment_description
        ]
        status_change_list_p_s_t['partner'] = [
            partners != partners_dataset,
            partners
        ]
        status_change_list_p_s_t['payment_due_date'] = [
            payment_due_date != payment_due_date_dataset,
            payment_due_date
        ]
        status_change_list_p_s_t['our_companies_id'] = [
            our_companies_id != our_companies_id_dataset,
            our_companies_id
        ]
        status_change_list_p_s_t['payment_sum'] = [
            payment_sum != payment_sum_dataset,
            payment_sum
        ]

        status_change_list_p_s_t['payment_full_agreed_status'] = [
            payment_full_agreed_status != payment_full_agreed_status_dataset,
            payment_full_agreed_status
        ]

        # True/False для определения, изменил ли пользователь согласованный остаток в payments_approval
        status_change_list_p_a = approval_sum != approval_sum_dataset

        print('status_change_list_p_a', status_change_list_p_a)

        conn, cursor = app_login.conn_cursor_init_dict()
        # Информация о платеже
        cursor.execute(
            """SELECT 
                t1.payment_id,
                t1.payment_sum::float,
                COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum)::float AS unapproved_sum,
                COALESCE(t2.approval_sum, 0)::float AS approval_sum,
                COALESCE(t9.approval_sum, 0)::float AS approval_to_pay_sum,
                COALESCE(t3.paid_sum, 0)::float AS paid_sum

        FROM public.payments_summary_tab AS t1
        LEFT JOIN (
                SELECT DISTINCT ON (payment_id) 
                    payment_id,
                    status_id,
                    SUM(approval_sum) OVER (PARTITION BY payment_id) AS approval_sum
                FROM public.payments_approval_history
                ORDER BY payment_id, created_at DESC
        ) AS t2 ON t1.payment_id = t2.payment_id
        LEFT JOIN (
                SELECT payment_id,
                    SUM(paid_sum) OVER (PARTITION BY payment_id) AS paid_sum
                FROM public.payments_paid_history
        ) AS t3 ON t1.payment_id = t3.payment_id

        LEFT JOIN (
                SELECT 
                    payment_id,
                    approval_sum
                FROM public.payments_approval
        ) AS t9 ON t1.payment_id = t9.payment_id
        WHERE t1.payment_id = %s
        ORDER BY t1.payment_due_date;
            """,
            [payment_id]
        )
        payment_data = cursor.fetchone()

        # Значения для обновления таблицы payments_approval
        values_p_a.append([
            payment_data['approval_to_pay_sum'] - payment_data['approval_sum'] + approval_sum,  # approval_sum
            payment_id  # payment_id
        ])

        ###########################################################################################
        # Проверка ДВУХ сумм (общая и согласованная)
        ###########################################################################################
        # Общая сумма
        if status_change_list_p_s_t['payment_sum'][0]:
            # Если была изменена общая сумма и она оказалась меньше общей оплаченной суммы => ОШИБКА
            if payment_data['paid_sum'] > status_change_list_p_s_t['payment_sum'][1]:
                app_login.conn_cursor_close(cursor, conn)
                description = (f"{'{0:,}'.format(status_change_list_p_s_t['payment_sum'][1]).replace(',', ' ')} ₽ > "
                               f"{'{0:,}'.format(payment_data['paid_sum']).replace(',', ' ')} ₽")
                flash(message=['ОШИБКА. Сумма платежа меньше оплаченной суммы', description], category='error')
                return jsonify({
                    'status': 'error',
                    'description': description,
                })
        else:
            # Если сумма не изменилась, то для удобства использования общей суммы в последующих расчетах заменяем сумму
            # в status_change_list_p_s_t['payment_sum'] на payment_data['payment_sum']
            status_change_list_p_s_t['payment_sum'][1] = payment_data['payment_sum']

        # Согласованная сумма
        if status_change_list_p_a:
            # Если согласованная сумма оказалась меньше общей оплаченной суммы => ОШИБКА
            if payment_data['paid_sum'] > approval_sum:
                app_login.conn_cursor_close(cursor, conn)
                description = (f"{'{0:,}'.format(approval_sum).replace(',', ' ')} ₽ < "
                               f"{'{0:,}'.format(payment_data['paid_sum']).replace(',', ' ')} ₽")
                flash(message=['ОШИБКА. Согласованная сумма платежа меньше оплаченной суммы', description],
                      category='error')
                return jsonify({
                    'status': 'error',
                    'description': description,
                })
            # Если согласованная сумма больше общей суммы => ОШИБКА
            elif approval_sum > status_change_list_p_s_t['payment_sum'][1]:
                app_login.conn_cursor_close(cursor, conn)
                description = (f"{'{0:,}'.format(approval_sum).replace(',', ' ')} ₽ > "
                               f"{'{0:,}'.format(status_change_list_p_s_t['payment_sum'][1]).replace(',', ' ')} ₽")
                flash(message=['ОШИБКА. Согласованная сумма платежа больше общей суммы', description], category='error')
                return jsonify({
                    'status': 'error',
                    'description': description,
                })
        else:
            approval_sum = payment_data['approval_sum']

        # Общая сумма меньше согласованной суммы
        if status_change_list_p_s_t['payment_sum'][0] and approval_sum > status_change_list_p_s_t['payment_sum'][1]:
            app_login.conn_cursor_close(cursor, conn)
            description = (f"{'{0:,}'.format(status_change_list_p_s_t['payment_sum'][1]).replace(',', ' ')} ₽ > "
                           f"{'{0:,}'.format(approval_sum).replace(',', ' ')} ₽")
            flash(message=['ОШИБКА. Сумма платежа меньше согласованной суммы', description], category='error')
            return jsonify({
                'status': 'error',
                'description': description,
            })
        # Остаток к оплате отрицательный (слишком сильно уменьшили сумму согласования)
        if values_p_a[0][0] < 0:
            app_login.conn_cursor_close(cursor, conn)
            description = (f"{'{0:,}'.format(values_p_a[0][0]).replace(',', ' ')} ₽")
            flash(message=['ОШИБКА. Остаток к оплате стал отрицательным', description], category='error')
            return jsonify({
                'status': 'error',
                'description': description,
            })

        """для db payments_summary_tab"""
        values_p_s_t.append([
            payment_id  # id заявки
        ])

        columns_p_s_t.append(
            "payment_id"
        )

        for k, v in status_change_list_p_s_t.items():
            if v[0]:
                values_p_s_t[0].append(v[1])
                columns_p_s_t.append(k)

        """для db payments_approval"""
        if values_p_a[0][0] or status_change_list_p_a:
        # if status_change_list_p_s_t['payment_sum'][0] or status_change_list_p_a:
            columns_p_a = ('approval_sum', 'payment_id')

        """для db payments_approval_history"""
        approval_sum_p_a_h = approval_sum - payment_data['approval_sum']  # Корректировка согласованной суммы
        values_p_a_h.append([
            payment_id,  # payment_id
            status_id,  # status_id
            user_id,  # user_id
            approval_sum_p_a_h,  # approval_sum
        ])

        columns_p_a_h = (
            "payment_id",
            "status_id",
            "user_id",
            "approval_sum"
        )

        columns_p_s_t = tuple(columns_p_s_t)

        # Изменяем запись в таблице payments_summary_tab
        if len(columns_p_s_t) > 1:
            query_p_s_t = get_db_dml_query(action='UPDATE', table='payments_summary_tab', columns=columns_p_s_t)

            execute_values(cursor, query_p_s_t, values_p_s_t)

            conn.commit()

            # Изменяем запись в таблице payments_approval_history
        if approval_sum_p_a_h:
            action_p_a_h = 'INSERT INTO'
            table_p_a_h = 'payments_approval_history'
            query_a_h = get_db_dml_query(action=action_p_a_h, table=table_p_a_h, columns=columns_p_a_h)

            execute_values(cursor, query_a_h, values_p_a_h)

            conn.commit()

        if columns_p_a:
            # Если сумма не равна нулю, обновляем значение - иначе удаляем строку из таблицы
            if values_p_a[0][0]:
                action_p_a = 'INSERT CONFLICT UPDATE'
                table_p_a = 'payments_approval'
                expr_set = ', '.join([f"{col} = EXCLUDED.{col}" for col in columns_p_a[:-1]])
                query_p_a = get_db_dml_query(
                    action=action_p_a, table=table_p_a, columns=columns_p_a, expr_set=expr_set
                )

            else:
                columns_p_a = 'payment_id'
                query_p_a = get_db_dml_query(action='DELETE', table='payments_approval', columns=columns_p_a)
                values_p_a = ((values_p_a[0][1],),)

            execute_values(cursor, query_p_a, values_p_a)

            conn.commit()

        app_login.conn_cursor_close(cursor, conn)

        if len(columns_p_s_t) > 1 or approval_sum_p_a_h:
            flash(message=['Заявка обновлена', ''], category='success')
            return jsonify({'status': 'success'})

        else:
            flash(message=['Изменения не найдены', 'Заявка на оплату не изменена'], category='error')
            return jsonify({'status': 'error'})

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return jsonify({'status': 'error'})


@payment_app_bp.route('/annul_payment', methods=['POST'])
@login_required
def annul_payment():
    """Аннулирование платежа из карточки платежа"""
    try:
        user_id = app_login.current_user.get_id()
        try:
            payment_number = int(request.get_json()['paymentId'])  # Номера платежей (передаётся id)
        except:
            payment_number = None
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, log_description=payment_number, user_id=user_id,
                               ip_address=app_login.get_client_ip())

        status_id = 6  # Статус заявки ("Аннулирован")
        values_p_s_t = []  # Данные для записи в таблицу payments_summary_tab
        values_p_a_h = []  # Данные для записи в таблицу payments_approval_history

        # Данные для удаления временных данных из таблицы payments_summary_tab
        values_p_d = []
        page_name = 'payment-approval'
        parameter_name = 'amount'

        values_a_h = []  # Список согласованных заявок для записи на БД

        values_a_h.append([
            payment_number,
            status_id,
            user_id
        ])

        conn, cursor = app_login.conn_cursor_init_dict()

        """проверяем, не закрыт ли платеж уже"""
        cursor.execute(
            """SELECT 
                    payment_close_status
            FROM public.payments_summary_tab
            WHERE payment_id = %s""",
            [payment_number]
        )
        payment_close_status = cursor.fetchone()[0]

        """Общая согласованная сумма"""
        cursor.execute(
            """SELECT
                    SUM(approval_sum) AS approval_sum
                FROM public.payments_approval_history
                WHERE payment_id = %s""",
            [payment_number]
        )
        approval_sum = cursor.fetchone()[0]

        if payment_close_status:
            flash(message=['Заявка не аннулирована', ''], category='error')
            return jsonify({'status': 'error'})

        """для db payments_summary_tab"""
        values_p_s_t.append([
            payment_number,  # id заявки
            True  # Закрытие заявки
        ])

        """для db payment_draft"""
        values_p_d.append((
            page_name,
            payment_number
        ))

        """для db payments_approval_history"""
        values_p_a_h.append([
            payment_number,
            status_id,
            user_id
        ])

        try:
            columns_p_s_t = ("payment_id", "payment_close_status")
            query_p_s_t = get_db_dml_query(action='UPDATE', table='payments_summary_tab', columns=columns_p_s_t)
            execute_values(cursor, query_p_s_t, values_p_s_t)

            columns_p_d = 'page_name, parent_id::int'
            query_p_d = get_db_dml_query(action='DELETE', table='payment_draft', columns=columns_p_d)
            execute_values(cursor, query_p_d, (values_p_d,))

            # Запись в payments_approval_history
            action_p_a_h = 'INSERT INTO'
            table_p_a_h = 'payments_approval_history'
            columns_p_a_h = ('payment_id', 'status_id', 'user_id')
            query_a_h = get_db_dml_query(
                action=action_p_a_h, table=table_p_a_h, columns=columns_p_a_h
            )
            execute_values(cursor, query_a_h, values_p_a_h)
            conn.commit()

            flash(message=['Заявка аннулирована', ''], category='success')

            app_login.conn_cursor_close(cursor, conn)
            return jsonify({'status': 'success'})
            # return redirect(url_for('.get_unapproved_payments'))

        except Exception as e:
            conn.rollback()
            app_login.conn_cursor_close(cursor, conn)
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning', flash_status=True)
            return jsonify({'status': 'error'})

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return jsonify({'status': 'error'})


@payment_app_bp.route('/annul_approval_payment', methods=['POST'])
@login_required
def annul_approval_payment():
    """Аннулирование согласованного из карточки платежа.
    В таблицу payments_approval_history добавляем запись со статусом 6 Аннулирован, и сумма,
    а из таблицы payments_approval удаляем всю согласованную сумму
    Для таблицы согласованных платежей - меняем статус закрытия заявки - на открытую,
    чтобы вернуть в список несогласованных платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        payment_number = int(request.get_json()['paymentId'])  # Номера платежей (передаётся id)
        page_url = request.get_json()['page_url']  # Страница, с которой вызвана функция
        status_id = 6  # Статус оплаты ("Аннулирован")
        page_name = 'payment-pay'

        conn, cursor = app_login.conn_cursor_init_dict()

        """Общая согласованная сумма"""
        cursor.execute(
            """SELECT
                    COALESCE(SUM(approval_sum), 0) AS approval_sum
                FROM public.payments_approval_history
                WHERE payment_id = %s""",
            [payment_number]
        )
        approval_sum = cursor.fetchone()[0]

        # Добавляем запись в таблицу payments_paid_history со статусом 11
        values_p_p_h = [(
            payment_number,
            11,
            user_id,
            0
        )]
        columns_p_p_h = ('payment_id', 'status_id', 'user_id', 'paid_sum')
        query_p_p_h = get_db_dml_query(action='INSERT INTO', table='payments_paid_history', columns=columns_p_p_h)
        execute_values(cursor, query_p_p_h, values_p_p_h)

        # Удаляем согласованную сумму из payments_approval
        values_a = [(payment_number,)]
        columns_a = 'payment_id'
        query_a = get_db_dml_query(action='DELETE', table='payments_approval', columns=columns_a)
        execute_values(cursor, query_a, (values_a,))

        # Удаляем временные данные из payment_draft
        values_p_d = [(
            page_name,
            payment_number
        )]
        columns_p_d = 'page_name, parent_id::int'
        query_p_d = get_db_dml_query(action='DELETE', table='payment_draft', columns=columns_p_d)
        execute_values(cursor, query_p_d, (values_p_d,))

        # Запись в payments_approval_history
        """для db payments_approval_history"""
        values_p_a_h = [
            (
                payment_number,
                status_id,
                user_id,
                -approval_sum
            ),
            (
                payment_number,
                4,
                user_id,
                0
            )
        ]
        table_p_a_h = 'payments_approval_history'
        columns_p_a_h = ('payment_id', 'status_id', 'user_id', 'approval_sum')
        query_a_h = get_db_dml_query(action='INSERT INTO', table=table_p_a_h, columns=columns_p_a_h)
        execute_values(cursor, query_a_h, values_p_a_h)

        conn.commit()

        if page_url == 'payment-approval-list':
            values_p_s_t = [(
                payment_number,  # id заявки
                False  # Открытие заявки
            )]
            columns_p_s_t = ("payment_id", "payment_close_status")
            query_p_s_t = get_db_dml_query(action='UPDATE', table='payments_summary_tab', columns=columns_p_s_t)
            execute_values(cursor, query_p_s_t, values_p_s_t)

        conn.commit()
        app_login.conn_cursor_close(cursor, conn)

        flash(message=['Согласования аннулированы', ''], category='success')
        return jsonify({'status': 'success'})

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return jsonify({
            'status': 'error',
            'description': msg_for_user,
        })


# Создание запроса в БД для множественного внесения данных
@login_required
def get_db_dml_query(action, table, columns, expr_set=None, subquery=";"):
    query = None
    if action == 'UPDATE':
        # В columns первым значением списка должна быть колонка для WHERE.
        # Связано с правилом выполнения sql-запроса

        # Список столбцов в SET
        expr_set = ', '.join([f"{col.split(':')[0]} = c.{col}" for col in columns[1:]])
        # Список столбцов для таблицы "с"
        col_with_out_type = tuple([i.split(':')[0] for i in columns])
        # col_with_out_type = tuple([i.split(':')[0] for i in columns[0]] + [i.split(':')[0] for i in columns[1:]])
        expr_s_tab = str(col_with_out_type).replace('\'', '').replace('"', '')
        # Выражение для WHERE
        expr_where = f"c.{columns[0]} = t.{columns[0]}"
        # Конструктор запроса
        query = f"{action} {table} AS t SET {expr_set} FROM (VALUES %s) AS c {expr_s_tab} WHERE {expr_where} {subquery}"

    elif action == 'UPDATE DOUBLE':
        # В columns первым значением списка должна быть колонка для WHERE. В первом значении передается список столбцов.

        # Список столбцов в SET
        expr_set = ', '.join([f"{col.split(':')[0]} = c.{col}" for col in columns[1:]])
        # Список столбцов для таблицы "с"
        col_with_out_type = tuple([i.split(':')[0] for i in columns[0]] + [i.split(':')[0] for i in columns[1:]])
        expr_s_tab = str(col_with_out_type).replace('\'', '').replace('"', '')
        # Выражение для WHERE
        expr_where = ' AND '.join([f"c.{col} = t.{col}" for col in columns[0]])
        # Конструктор запроса
        query = f"UPDATE {table} AS t SET {expr_set} FROM (VALUES %s) AS c {expr_s_tab} WHERE {expr_where} {subquery}"

    elif action == 'INSERT INTO':
        # Кортеж колонок переводим в строки и удаляем кавычки
        expr_cols = str(columns).replace('\'', '').replace('"', '')
        # Конструктор запроса
        query = f"{action} {table} {expr_cols} VALUES %s {subquery}"

    elif action == 'INSERT CONFLICT UPDATE':
        # В columns последним значением списка должна быть колонка для ON CONFLICT.
        # Связано с тем, что из value мы берем значения по порядку, все кроме ON CONFLICT колонки

        # Кортеж колонок переводим в строки и удаляем кавычки
        expr_cols = str(columns).replace('\'', '').replace('"', '')
        # Конструктор запроса
        query = f"INSERT INTO {table} AS t1 {expr_cols} VALUES %s ON CONFLICT ({columns[-1]}) DO UPDATE SET {expr_set};"

    elif action == 'DELETE':
        query = f"DELETE FROM {table} WHERE ({columns}) IN %s;"

    return query


# Превращаем строковое значение стоимости с пропусками и ₽ в число
def convert_amount(amount):
    try:
        amount = float(amount.replace('₽', '').replace(" ", "").replace(" ", "").replace(",", "."))
    except:
        amount = None
    return amount


# Функция преобразовывает значение в запросе для WHERE при пагинации
def conv_data_to_db(col, val, all_col_types, manual_type='', sign=''):
    # Если указан ручной тип данных, то не ищем тип данных из БД
    if manual_type:
        col_type = manual_type
    else:
        # Преобразовываем название колонки. Название может быть 't1.payment_id', нам нужно убрать назв таблицы (t1)
        col = col.split('.')[1] if len(col.split('.')) > 1 else col
        # Тип данных таблицы
        col_type = ''
        for lst in all_col_types:
            if lst[0] == col:
                col_type = lst[1]
                break

    # В зависимости от типа колонки преобразовываем значение и указываем тип
    # Числа
    if (col_type == 'real' or
            col_type == 'bigint' or
            col_type == 'smallint' or
            col_type == 'numeric' or
            col_type == 'integer'):
        val = f"{val}::{col_type}"
    # даты
    elif col_type == 'date' or col_type == 'timestamp with time zone' or col_type == 'timestamp without time zone':
        val = f"timestamp without time zone '{val}'"
        # if val.lower() == 'null' or not val:
        #     val = f"'{sign}infinity'::timestamp without time zone"
        # else:
        #     val = f"timestamp without time zone '{val}'"
    # Текст
    elif col_type == 'text' or col_type == 'character varying':
        val = f"'{replace_percent(val)}'::text"
    # boolean
    elif col_type == 'boolean':
        pass
    else:
        val = f"'{replace_percent(val)}'::text"
    return val


@payment_app_bp.route('/get-paymentMyCharts', methods=['POST'])
@login_required
def get_payment_my_charts():
    """График платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        chart_type = request.get_json()['chart_type']

        query_tab = ''
        label = ''
        title = ''

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        # Данные для графика изменения доступного к распределению
        if chart_type == 'available_money':
            query_tab = 't22'
            title = 'История изменения доступных к распределению средств'
            label = 'Доступно к распределению'

        # Данные для графика изменения баланса
        if chart_type == 'account_money':
            query_tab = 't11'
            title = 'История изменения баланса на счету'
            label = 'Средства на счету'

        cursor.execute(
            f"""WITH
                t1 AS (SELECT 
                        sum(inflow_sum) AS balance_sum,
                        inflow_at AS created_at,
                        'payments_inflow_history' AS description,
                        'inflow' AS status
                    FROM public.payments_inflow_history
                    WHERE inflow_type_id != 4 AND inflow_sum != 0
                    GROUP BY inflow_at
                    ORDER BY inflow_at DESC
                    LIMIT 40
                ),
                t2 AS (SELECT 
                        sum(paid_sum)*-1 AS balance_sum,
                        created_at AS created_at,
                        'payments_paid_history' AS description,
                        'paid' AS status
                    FROM public.payments_paid_history
                    WHERE paid_sum != 0
                    GROUP BY created_at
                    ORDER BY created_at DESC
                    LIMIT 40
                ),
                t3 AS (SELECT 
                        sum(approval_sum)*-1 AS balance_sum,
                        created_at AS created_at,
                        'payments_approval_history' AS description,
                        'approval' AS status
                    FROM public.payments_approval_history
                    WHERE status_id != 1 AND approval_sum != 0
                    GROUP BY created_at
                    ORDER BY created_at DESC
                    LIMIT 40
                ),
                t4 AS (SELECT 
                        COALESCE(SUM(balance_sum), 0) AS all_sum
                    FROM public.payments_balance
                ),
                    t5 AS (SELECT 
                            COALESCE(SUM(approval_sum), 0) AS approval_sum
                        FROM public.payments_approval
                    ),
                t11 AS (
                    SELECT
                        *
                    FROM t1
                    JOIN t4 ON true
                    UNION ALL 
                    SELECT
                        *
                    FROM t2
                    JOIN t4 ON true
                    ORDER BY created_at DESC
                    LIMIT 40
                ),
                    t22 AS (
                        SELECT 
                            t1.created_at,
                            (t4.all_sum - t5.approval_sum) AS all_sum,
                            t1.balance_sum,
                            t1.status
                        FROM t1
                        JOIN t4 ON true
                        JOIN t5 ON true
                        UNION ALL 
                        SELECT
                            t3.created_at,
                            (t4.all_sum - t5.approval_sum) AS all_sum,
                            t3.balance_sum,
                            t3.status
                        FROM t3
                        JOIN t4 ON true
                        JOIN t5 ON true
                        ORDER BY created_at DESC
                        LIMIT 40
                    ) 
                SELECT 
                    date_trunc('second', created_at::timestamp without time zone)::text AS created_at,
                    COALESCE(all_sum - SUM(balance_sum) OVER (ORDER BY created_at DESC) + balance_sum, all_sum)::text AS cur_bal,
                    status
                FROM {query_tab}
                ORDER BY created_at;
            """
        )
        historic_data = cursor.fetchall()

        if not len(historic_data):
            return jsonify({
                'status': 'error',
                'description': 'Данные не найдены',
            })

        for i in range(len(historic_data)):
            historic_data[i] = dict(historic_data[i])

        app_login.conn_cursor_close(cursor, conn)

        # Return the updated data as a response
        return jsonify({
            'historic_data': historic_data,
            'title': title,
            'label': label,
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'historic_data': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/payment-paid-list-for-a-period', methods=['GET'])
@login_required
def payments_paid_list_for_a_period():
    """Выгрузка из БД списка оплаченных платежей за период"""
    try:
        global hlink_menu, hlink_profile

        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())
        role = app_login.current_user.get_role()

        # Check if the user has access to the "List of contracts" page
        if role not in (1, 4, 6):
            return error_handlers.handle403(403)

        # Create profile name dict
        hlink_menu, hlink_profile = app_login.func_hlink_profile()

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=request.path[1:])

        tab_rows = 1
        flash(message=[
            'Краткое описание', '',
            'Есть 3 кнопки - выгрузки за 1 (текущий день), за последние 7 дней и за 30 дней',
            'Есть кнопка "Период" и два поля периода. '
            'Если указать только одну дату (в любом поле из двух) - получим данные за указанную дату. '
            'Если указаны две даты (не имеет значение какая дата больше другой), то отобразится информация '
            'за выбранный период (даты включены в выборку)',
            'Используя фильтры в таблице можно найти что нужно и нажать кнопку "Обновить"'
        ], category='success')


        # Список разделов в меню платежей
        payment_approval_menu = get_payment_approval_menu(role)

        return render_template('payment-paid-list-card_data_for_a_period.html', menu=hlink_menu,
                               menu_profile=hlink_profile, sort_col='sort_col', tab_rows=tab_rows, nonce=get_nonce(),
                               setting_users=setting_users, payment_approval_menu=payment_approval_menu,
                               title='Оплаченные платежи за период')
    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return render_template('page_error.html', error=['Ошибка', msg_for_user], nonce=get_nonce())


@payment_app_bp.route('/get-payment_paid_data-for-a-period', methods=['POST'])
@login_required
def get_payment_paid_data_for_a_period():
    """Выгрузка списка оплаченных платежей за указанный период"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_name = 'payment-paid-list-for-a-period'
        date_first = request.get_json()['dateFirst']
        date_second = request.get_json()['dateSecond']
        col_1 = request.get_json()['sort_col_1']
        col_1_val = request.get_json()['sort_col_1_val']
        col_id = 't1.payment_id'
        col_id_val = request.get_json()['sort_col_id_val']
        filter_vals_list = request.get_json()['filterValsList']

        if not date_first or not date_second or len(date_first) != 10 or len(date_second) != 10:
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': ['Ошибка при получение даты. rev-1'],
            })

        date_first = date.fromisoformat(date_first)
        date_second = date.fromisoformat(date_second)

        limit = 0

        # Выражение для фильтрации в выражении WHERE
        where_expression = f"t0.paid_sum > 0 AND (t0.created_at::date BETWEEN SYMMETRIC %s AND %s) "
        where_expression2 = []  # Вторая часть условия
        query_value = [date_second, date_first]
        # Колонка по которой идёт сортировка в таблице
        col_num = int(col_1.split('#')[0])
        # Направление сортировки
        sort_direction = col_1.split('#')[1]

        # Список колонок для сортировки
        sort_col = {
            'col_1': [f"{col_num}#{sort_direction}"],  # Первая колонка
            'col_id': ''
        }

        # столбцы фильтров
        col_0 = ""
        col_1 = "t1.payment_number"
        col_2 = "t4.cost_item_name"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "t6.object_name"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "t0.paid_sum"
        col_10 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_11 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_12 = "to_char(t0.created_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_13 = "t8.status_name"
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]
        # столбцы сортировки
        col_0 = ""
        col_1 = "t1.payment_number"
        col_2 = "t4.cost_item_name"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "COALESCE(t6.object_name, '')"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "t0.paid_sum"
        col_10 = "t1.payment_due_date"
        col_11 = "t1.payment_at"
        col_12 = "t0.created_at"
        col_13 = "t8.status_name"
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]
        sort_col_1, sort_col_1_order = list_sort_col[col_num], 'DESC' if sort_direction == '1' else 'ASC'
        sort_col_id, sort_col_id_order = col_id, 'DESC' if sort_direction == '1' else 'ASC'

        if filter_vals_list:
            for i in filter_vals_list:
                query_value.append('%' + i[1] + '%')
                where_expression2.append(list_filter_col[i[0]])
            where_expression2 = ' AND '.join(map(lambda x: f'{x}::text ILIKE %s', where_expression2))
        if where_expression2:
            where_expression += ' AND ' + where_expression2
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        try:
            cursor.execute(
                f"""
                SELECT 
                    t0.payment_id,
                    t1.payment_number,
                    t4.cost_item_name,
                    t1.basis_of_payment,
                    SUBSTRING(t1.basis_of_payment, 1,70) AS basis_of_payment_short,  
                    t3.contractor_name,
                    SUBSTRING(t1.payment_description, 1,70) AS payment_description_short,
                    t1.payment_description,
                    COALESCE(t6.object_name, '') AS object_name,
                    t5.first_name,
                    t5.last_name,
                    t1.partner,  
                    t1.payment_sum,
                    TRIM(BOTH ' ' FROM to_char(t1.payment_sum, '999 999 990D99 ₽')) AS payment_sum_rub,
                    t0.paid_sum AS paid_sum,
                    TRIM(BOTH ' ' FROM to_char(COALESCE(t0.paid_sum, 0), '999 999 990D99 ₽')) AS paid_sum_rub,
                    t1.payment_due_date::text AS payment_due_date,
                    to_char(t1.payment_due_date, 'dd.mm.yyyy') AS payment_due_date_txt,
                    t1.payment_at::timestamp without time zone::text AS payment_at,
                    to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS payment_at_txt,
                    t0.created_at AS paid_at,
                    to_char(t0.created_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS') AS paid_at_txt,
                    t8.status_name,
                    TRIM(BOTH ' ' FROM to_char(SUM(t0.paid_sum) OVER (), '999 999 990D99 ₽')) AS card_summary_paid_value
                    
                FROM public.payments_paid_history AS t0
                LEFT JOIN (
                    SELECT 
                        payment_id, 
                        payment_number, 
                        basis_of_payment,
                        payment_description,
                        partner,
                        payment_sum,
                        payment_due_date,
                        payment_at,
                        our_companies_id,
                        cost_item_id,
                        responsible,
                        object_id
                    FROM public.payments_summary_tab
                ) AS t1 ON t0.payment_id = t1.payment_id
                LEFT JOIN (
                    SELECT contractor_id,
                        contractor_name
                    FROM public.our_companies            
                ) AS t3 ON t1.our_companies_id = t3.contractor_id
                LEFT JOIN (
                    SELECT cost_item_id,
                        cost_item_name
                    FROM public.payment_cost_items            
                ) AS t4 ON t1.cost_item_id = t4.cost_item_id
                LEFT JOIN (
                        SELECT user_id,
                            first_name,
                            last_name
                        FROM public.users
                ) AS t5 ON t1.responsible = t5.user_id
                LEFT JOIN (
                        SELECT object_id,
                            object_name
                        FROM public.objects
                ) AS t6 ON t1.object_id = t6.object_id
                LEFT JOIN (
                        SELECT payment_agreed_status_id AS status_id,
                            payment_agreed_status_name AS status_name
                        FROM public.payment_agreed_statuses
                ) AS t8 ON t0.status_id = t8.status_id
                WHERE {where_expression}
                ORDER BY {sort_col_1} {sort_col_1_order}, {sort_col_id} {sort_col_id_order};
            """,
                query_value
            )
            all_payments = cursor.fetchall()

        except Exception as e:
            msg_for_user = app_login.create_traceback(info=sys.exc_info(), error_type='warning')
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': 0,
                'sort_col': 0,
                'status': 'error',
                'description': ['Ошибка при выгрузки данных. rev-1' + msg_for_user],
            })
        if not len(all_payments):
            app_login.conn_cursor_close(cursor, conn)
            return jsonify({
                'payment': None,
                'sort_col': sort_col,
                'status': 'success',
                'description': ['Конец таблицы. Ничего не найдено'],
            })
        else:
            for i in range(len(all_payments)):
                all_payments[i] = dict(all_payments[i])

        app_login.conn_cursor_close(cursor, conn)

        # Настройки таблицы
        setting_users = get_tab_settings(user_id=user_id, list_name=page_name)

        # Return the updated data as a response
        return jsonify({
            'payment': all_payments,
            'sort_col': sort_col,
            'tab_rows': len(all_payments),
            'page': page_name,
            'setting_users': setting_users,
            'direction': 'down' if sort_direction == '1' else 'up',
            'status': 'success',
            'description': ["Данные загружены", f"Выгружены данные за {date_second} - {date_first}",
                            f"Всего найдено платежей: {len(all_payments)}"]
        })

    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


@payment_app_bp.route('/save_tab_settings', methods=['POST'])
@login_required
def save_tab_settings():
    """Сохранение изменений отображаемых полей пользователя на различных страницах"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id, ip_address=app_login.get_client_ip())

        page_url = request.get_json()['page_url']
        hide_list = request.get_json()['hide_list']

        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        # Список скрываемых столбцов пользователя
        cursor.execute(
            f"""
            SELECT
                 unit_name
            FROM setting_users
            WHERE user_id = %s AND list_name = %s
            """,
            [user_id, page_url]
        )

        hide_list_db = cursor.fetchall()
        hide_list_db = [x[0] for x in hide_list_db]
        hide_list = [str(x) for x in hide_list]

        show_list = [x for x in hide_list if x not in hide_list_db]
        hide_list = [x for x in hide_list_db if x not in hide_list]

        if len(hide_list):
            columns_del = 'user_id::int, list_name, unit_name'
            query_del = get_db_dml_query(action='DELETE', table='setting_users', columns=columns_del)
            values_del = [(user_id, page_url, str(i)) for i in hide_list]

            execute_values(cursor, query_del, (values_del,))

        if len(show_list):
            columns_ins = ('user_id', 'list_name', 'unit_name')
            values_ins = [(user_id, page_url, str(i)) for i in show_list]
            query_ins = get_db_dml_query(action='INSERT INTO', table='setting_users', columns=columns_ins)
            execute_values(cursor, query_ins, values_ins)

        conn.commit()
        app_login.conn_cursor_close(cursor, conn)

        # Return the updated data as a response
        return jsonify({
            'tab_rows': 'tab_rows',
            'page': 'payment-pay',
            'status': 'success'
        })
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return jsonify({
            'payment': 0,
            'sort_col': 0,
            'status': 'error',
            'description': msg_for_user,
        })


def get_tab_settings(user_id=0, list_name=0, unit_name=0, unit_value=0):
    """Список отображаемых полей пользователя на различных страницах"""
    try:
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        # Список скрываемых столбцов пользователя
        cursor.execute(
            f"""
                   SELECT
                        unit_name,
                        unit_value
                   FROM setting_users
                   WHERE user_id = %s AND list_name = %s
                   ORDER BY unit_name DESC
                   """,
            [user_id, list_name]
        )

        setting_users = cursor.fetchall()

        app_login.conn_cursor_close(cursor, conn)

        setting_users2 = dict()

        if len(setting_users):
            for i in range(len(setting_users)):
                setting_users2[str(setting_users[i][0])] = (True, setting_users[i]['unit_value'])
                setting_users[i] = dict(setting_users[i])

        return setting_users2

    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return False


# Скрываем несогласованные заявки на оплату со страницы список платежей
@payment_app_bp.route('/hide_payment', methods=['POST'])
@login_required
def hide_payment():
    """Скрываем несогласованные заявки на оплату со страницы список платежей"""
    try:
        user_id = app_login.current_user.get_id()
        app_login.set_info_log(log_url=sys._getframe().f_code.co_name, user_id=user_id,
                               ip_address=app_login.get_client_ip())

        payment_id: int = int(request.get_json()['payment_id'])

        conn, cursor = app_login.conn_cursor_init_dict()

        # Проверяем, что у платежа не было никаких изменений в таблице согласования
        cursor.execute(
            """
            SELECT
                COUNT(*)
            FROM public.payments_approval_history
            WHERE payment_id = %s
            GROUP BY payment_id ;""",
            [payment_id]
        )
        delete_status = cursor.fetchone()[0]
        delete_status = True if delete_status == 1 else False

        if delete_status:
            # Определяем, что пользователь был создателем или ответственным.
            #Если ответственным, то меняем только ответственного и сообщаем, что задача не была скрыта полностью,
            # а только скрыта у него
            cursor.execute(
                """
                SELECT
                    payment_owner,
                    responsible
                FROM public.payments_summary_tab
                WHERE payment_id = %s ;""",
                [payment_id]
            )
            own_resp = cursor.fetchall()
            payment_owner, responsible = (own_resp[0][0], own_resp[0][1]) if own_resp else (0, 0)

            # Пользователь является и ответственным и создателем задачи
            if payment_owner == responsible == user_id:
                # Обновляем описание и пользователя в задаче
                query_upd = """
                UPDATE payments_summary_tab 
                SET 
                    payment_description = (
                        SELECT 
                            concat_ws(' ', 
                                'owner:', 
                                payment_owner::text, 
                                'responsible:', 
                                responsible::text, 
                                'description:', 
                                payment_description) 
                            FROM public.payments_summary_tab 
                            WHERE payment_id = %s),
                    payment_owner = 1,
                    responsible = 1,
                    payment_close_status = TRUE
                WHERE payment_id = %s;
                """
                values_upd = [payment_id, payment_id]
                cursor.execute(query_upd, values_upd)

                # Добавляем запись об аннулировании
                columns_ins = ('payment_id', 'status_id', 'user_id')
                values_ins = [[payment_id, 6, 1]]
                query_ins = get_db_dml_query(action='INSERT INTO', table='payments_approval_history',
                                                            columns=columns_ins)
                execute_values(cursor, query_ins, values_ins)

                conn.commit()
                app_login.conn_cursor_close(cursor, conn)

                flash(message=[f'Платеж №{payment_id} аннулирован и скрыт'], category='success')
                return jsonify({
                    'status': 'success',
                    'description': ['Платеж аннулирован и скрыт'],
                })
            # Пользователь только ответственный
            elif responsible == user_id != payment_owner:
                # Обновляем описание и пользователя в задаче
                query_upd = """
                UPDATE payments_summary_tab 
                SET 
                    responsible = 1
                WHERE payment_id = %s;
                """
                values_upd = [payment_id]
                execute_values(cursor, query_upd, values_upd)

                conn.commit()
                app_login.conn_cursor_close(cursor, conn)

                flash(message=[f'Платеж №{payment_id} скрыт',
                               'Платеж не аннулирован, т.к. вы не являетесь создателем заявки'], category='success')
                return jsonify({
                    'status': 'success',
                    'description': ['Платеж скрыт, не аннулирован'],
                })
            # В остальных случаях - ОШИБКА
            app_login.conn_cursor_close(cursor, conn)
            flash(message=['Ошибка', f'Платеж №{payment_id} не скрыт'], category='error')
            return jsonify({
                'status': 'error',
                'description': ['Платеж не скрыт'],
            })
        # UPDATE payments_summary_tab SET payment_close_status = false WHERE payment_id = 313;
        else:
            flash(message=['Ошибка', f'Платеж №{payment_id} не удалось удалить',
                           'По платежу уже происходили действия по согласованию'], category='error')
            return jsonify({
                'status': 'error',
                'description': ['По платежу уже происходили действия по согласованию'],
            })

    except Exception as e:
        msg_for_user = app_login.create_traceback(info=sys.exc_info(), flash_status=True)
        return jsonify({
            'status': 'error',
            'description': [msg_for_user],
        })

# Получаем типы данных из всех столбцов всех таблиц БД
def get_table_list():
    try:
        # Connect to the database
        conn, cursor = app_login.conn_cursor_init_dict()

        # Список таблиц в базе данных и их типы
        cursor.execute(
            """
            SELECT column_name, data_type, table_name FROM information_schema.columns
            WHERE table_schema = 'public'
            ORDER BY column_name
            """
        )

        all_col_types = cursor.fetchall()
        app_login.conn_cursor_close(cursor, conn)

        return all_col_types
    except Exception as e:
        msg_for_user = app_login.create_traceback(sys.exc_info())
        return False


def get_payment_approval_menu(role: int):
    payment_approval_menu = []
    if role in (1, 4, 6):
        payment_approval_menu = [
            {'href': 'payment-approval', 'name': 'Согласование'},
            {'href': 'payment-approval-list', 'name': 'Согласованные'},
            {'href': 'payment-paid-list', 'name': 'Оплаченные'},
            {'href': 'payment-inflow-history-list', 'name': 'Входящие платежи'},
        ]
    elif role == 8:
        payment_approval_menu = []
    return payment_approval_menu


def get_sort_filter_data(page_name, limit, col_1, col_1_val, col_id, col_id_val, filter_vals_list, user_id,
                         manual_type=''):
    # Колонка по которой идёт сортировка в таблице
    col_num = int(col_1.split('#')[0])
    # Направление сортировки
    sort_direction = col_1.split('#')[1]

    # Список колонок для сортировки
    sort_col = {
        'col_1': [f"{col_num}#{sort_direction}"],  # Первая колонка
        'col_id': ''
    }

    query_value = [
        page_name,
        'amount',
        user_id
    ]

    if page_name == 'payment-approval':
        # столбцы фильтров
        col_0 = ""
        col_1 = """concat_ws(': ', 
                concat_ws(', ', t3.contractor_name, t6.object_name, CASE WHEN t1.partner<>'' THEN t1.partner END),
                concat_ws(' - ', t1.basis_of_payment, t1.payment_description))"""
        col_2 = "t1.payment_sum"
        col_3 = "COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum)"
        col_4 = "COALESCE(t8.amount, '0')"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_7 = "t7.status_name"
        col_8 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_9 = ""
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
        ]
        # столбцы сортировки
        col_0 = ""
        col_1 = """concat_ws(': ', 
                concat_ws(', ', t3.contractor_name, t6.object_name, CASE WHEN t1.partner<>'' THEN t1.partner END),
                concat_ws(' - ', t1.basis_of_payment, t1.payment_description))"""
        col_2 = "t1.payment_sum"
        col_3 = "COALESCE(t1.payment_sum - t2.approval_sum, t1.payment_sum)"
        col_4 = "COALESCE(t8.amount, '0')"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "t1.payment_due_date"
        col_7 = "t7.status_name"
        col_8 = "t1.payment_at"
        col_9 = ""
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
        ]
        # типы данных столбцов
        col_0 = ""
        col_1 = "t1.basis_of_payment"
        col_2 = "t1.payment_sum"
        col_3 = "t1.payment_sum"
        col_4 = "t8.approval_sum"
        col_5 = "t5.last_name"
        col_6 = "t1.payment_due_date"
        col_7 = "t7.status_name"
        col_8 = "t1.payment_at"
        col_9 = ""
        list_type_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
        ]
    elif page_name == 'payment-approval-list':
        # столбцы фильтров
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_4 = "t6.object_name"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "t0.approval_sum"
        col_9 = "COALESCE(t7.paid_sum, '0')"
        col_10 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_11 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_12 = "to_char(t8.created_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12
        ]
        # столбцы сортировки
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_4 = "COALESCE(t6.object_name, '')"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "t0.approval_sum"
        col_9 = "COALESCE(t7.paid_sum, '0')"
        col_10 = "t1.payment_due_date"
        col_11 = "t1.payment_at"
        col_12 = "t8.created_at"
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12
        ]
        # типы данных столбцов
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "t1.payment_description"
        col_4 = "t6.object_name"
        col_5 = "t5.last_name"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "t0.approval_sum"
        col_9 = "t7.paid_sum"
        col_10 = "t1.payment_due_date"
        col_11 = "t1.payment_at"
        col_12 = "t8.created_at"
        list_type_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12
        ]
        query_value = []
    elif page_name == 'payment-paid-list':
        # столбцы фильтров
        col_0 = ""
        col_1 = "t1.payment_number"
        col_2 = "t4.cost_item_name"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "t6.object_name"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "COALESCE(t2.approval_sum, '0')"
        col_10 = "t0.paid_sum"
        col_11 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_12 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_13 = "t8.status_name"
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]
        # столбцы сортировки
        col_0 = ""
        col_1 = "t1.payment_number"
        col_2 = "t4.cost_item_name"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "COALESCE(t6.object_name, '')"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "COALESCE(t2.approval_sum, '0')"
        col_10 = "t0.paid_sum"
        col_11 = "t1.payment_due_date"
        col_12 = "t1.payment_at"
        col_13 = "t8.status_name"
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]
        # типы данных столбцов
        col_0 = ""
        col_1 = "t1.payment_number"
        col_2 = "t4.cost_item_name"
        col_3 = "t1.basis_of_payment"
        col_4 = "t1.payment_description"
        col_5 = "t6.object_name"
        col_6 = "t5.last_name"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "t2.approval_sum"
        col_10 = "t0.paid_sum"
        col_11 = "t1.payment_due_date"
        col_12 = "t1.payment_at"
        col_13 = "t8.status_name"
        list_type_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13
        ]
        query_value = []
    elif page_name == 'payment-list':
        # столбцы фильтров
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "concat_ws(': ',  t3.contractor_name, t1.payment_description)"
        col_4 = "t6.object_name"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "COALESCE(t7.paid_sum, '0')"
        col_9 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_10 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10
        ]
        # столбцы сортировки
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "concat_ws(': ',  t3.contractor_name, t1.payment_description)"
        col_4 = "COALESCE(t6.object_name, '')"
        col_5 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "COALESCE(t7.paid_sum, '0')"
        col_9 = "t1.payment_due_date"
        col_10 = "t1.payment_at"
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10
        ]
        # типы данных столбцов
        col_0 = "t1.payment_number"
        col_1 = "t4.cost_item_name"
        col_2 = "t1.basis_of_payment"
        col_3 = "t1.payment_description"
        col_4 = "t6.object_name"
        col_5 = "t5.last_name"
        col_6 = "t1.partner"
        col_7 = "t1.payment_sum"
        col_8 = "t7.paid_sum"
        col_9 = "t1.payment_due_date"
        col_10 = "t1.payment_at"
        list_type_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10
        ]

        query_value = [
            user_id,
            user_id
        ]
    elif page_name == 'payment-pay':
        # столбцы фильтров
        col_0 = ""
        col_1 = "t4.cost_item_name"
        col_2 = "t1.payment_number"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "t6.object_name"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "COALESCE(t7.paid_sum, '0')"
        col_10 = "t0.approval_sum"
        col_11 = "COALESCE(t8.amount, '0')"
        col_12 = "to_char(t1.payment_due_date, 'dd.mm.yyyy')"
        col_13 = "to_char(t1.payment_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        col_14 = ""
        list_filter_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14
        ]
        # столбцы сортировки
        col_0 = ""
        col_1 = "t4.cost_item_name"
        col_2 = "t1.payment_number"
        col_3 = "t1.basis_of_payment"
        col_4 = "concat_ws(': ', t3.contractor_name, t1.payment_description)"
        col_5 = "COALESCE(t6.object_name, '')"
        col_6 = "concat_ws(' ', t5.last_name, t5.first_name)"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "COALESCE(t7.paid_sum, '0')::numeric"
        col_10 = "t0.approval_sum"
        col_11 = "COALESCE(t8.amount, '0')::numeric"
        col_12 = "t1.payment_due_date"
        col_13 = "t1.payment_at"
        col_14 = ""
        list_sort_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14
        ]
        # типы данных столбцов
        col_0 = ""
        col_1 = "t4.cost_item_name"
        col_2 = "t1.payment_number"
        col_3 = "t1.basis_of_payment"
        col_4 = "t1.payment_description"
        col_5 = "t6.object_name"
        col_6 = "t5.last_name"
        col_7 = "t1.partner"
        col_8 = "t1.payment_sum"
        col_9 = "t7.paid_sum"
        col_10 = "t0.approval_sum"
        col_11 = "t8.approval_sum"
        col_12 = "t1.payment_due_date"
        col_13 = "t1.payment_at"
        col_14 = ""
        list_type_col = [
            col_0, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14
        ]
    elif page_name == 'payment-inflow-history-list':
        # столбцы фильтров
        col_0 = "t1.inflow_id"
        col_1 = "t2.contractor_name"
        col_2 = "t1.inflow_description"
        col_3 = "t3.inflow_type_name"
        col_4 = "t1.inflow_sum"
        col_5 = "to_char(t1.inflow_at::timestamp without time zone, 'dd.mm.yyyy HH24:MI:SS')"
        list_filter_col = [col_0, col_1, col_2, col_3, col_4, col_5]
        # столбцы сортировки
        col_0 = "t1.inflow_id"
        col_1 = "t2.contractor_name"
        col_2 = "t1.inflow_description"
        col_3 = "t3.inflow_type_name"
        col_4 = "t1.inflow_sum"
        col_5 = "t1.inflow_at"
        list_sort_col = [col_0, col_1, col_2, col_3, col_4, col_5]
        # типы данных столбцов
        col_0 = "t1.inflow_id"
        col_1 = "t2.contractor_name"
        col_2 = "t1.inflow_description"
        col_3 = "t3.inflow_type_name"
        col_4 = "t1.inflow_sum"
        col_5 = "t1.inflow_at"
        list_type_col = [col_0, col_1, col_2, col_3, col_4, col_5]
        query_value = []

    sort_col_1, sort_col_1_order = list_sort_col[col_num], 'DESC' if sort_direction == '1' else 'ASC'
    sort_col_id, sort_col_id_order = col_id, 'DESC' if sort_direction == '1' else 'ASC'
    sort_col_1_equal = '>' if sort_col_1_order == 'ASC' else '<'

    # Список таблиц в базе данных и их типы
    all_col_types = get_table_list()
    # Выражение для фильтрации в выражении WHERE
    where_expression = (
        f"({sort_col_1}, {sort_col_id}) {sort_col_1_equal} "
        f"({conv_data_to_db(list_type_col[col_num], col_1_val, all_col_types)}, "
        f"{conv_data_to_db(sort_col_id, col_id_val, all_col_types)})")
    where_expression2 = []  # Вторая часть условия (пригодится для определения общего кол-ва строк)
    if filter_vals_list:
        for i in filter_vals_list:
            query_value.append('%' + replace_percent_slash(i[1]) + '%')
            where_expression2.append(list_filter_col[i[0]])
    where_expression2 = ' AND '.join(map(lambda x: f'{x}::text ILIKE %s', where_expression2))
    if where_expression2:
        where_expression += ' AND ' + where_expression2
    return sort_col_1, sort_col_1_order, sort_col_id, sort_col_id_order, where_expression, where_expression2, \
        query_value, sort_col, col_num

def replace_percent(text):
    # Use re.sub() to find and replace single % with %%
    return re.sub(r'(?<!%)%(?!%)', '%%', text)
def replace_percent_slash(text):
    # Use re.sub() to find and replace single % with %%
    return re.sub(r'(?<!%)%(?!%)', '\%', text)
